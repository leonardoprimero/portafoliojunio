#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Análisis Financiero Final - TODOS LOS PROBLEMAS CORREGIDOS
=========================================================
Análisis financiero profesional desarrollado por Leonardo Caliva.
Versión final con gráficos que NO se superponen con pie de página
y Excel con imágenes REALMENTE separadas de las tablas.

Autor: Leonardo Caliva
Portfolio: leocaliva.com
"""

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from matplotlib.backends.backend_pdf import PdfPages
import shutil
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings('ignore')

# ========================= CONFIGURACIÓN =========================

# CONFIGURACIÓN PRINCIPAL - CAMBIAR AQUÍ PARA USAR ESCALA LOGARÍTMICA
USAR_ESCALA_LOGARITMICA = False  # Cambiar a True para gráficos logarítmicos

# Configuración de archivos y carpetas
CARPETA_DATOS = './datospython1'  # Carpeta donde están los datos reales
CARPETA_GRAFICOS = './graficos_temp'
ARCHIVO_PDF = 'reporte_analisis_final.pdf'
ARCHIVO_EXCEL = 'analisis_activos_final.xlsx'

# Información del autor
AUTOR_NOMBRE = "Leonardo Caliva"
AUTOR_PORTFOLIO = "leocaliva.com"

# Configuración de matplotlib para A4 y sin superposición
def configurar_matplotlib():
    """Configurar matplotlib para A4 perfecto"""
    plt.style.use('seaborn-v0_8-whitegrid')
    sns.set_palette("Set2")
    plt.rcParams.update({
        'figure.figsize': (7.5, 5),  # MÁS PEQUEÑO para A4
        'axes.labelsize': 9,
        'axes.titlesize': 11,
        'legend.fontsize': 8,
        'xtick.labelsize': 8,
        'ytick.labelsize': 8,
        'font.size': 9,
        'pdf.fonttype': 42,
        'ps.fonttype': 42
    })

# ========================= GENERACIÓN DE DATOS DE PRUEBA =========================

def generar_datos_prueba():
    """Genera datos de prueba aleatorios para testing"""
    print("Generando datos de prueba aleatorios...")
    
    os.makedirs(CARPETA_DATOS, exist_ok=True)
    
    # Lista de activos de ejemplo
    activos = ['AAPL', 'MSFT', 'GOOGL', 'TSLA', 'AMZN']
    
    # Generar datos para cada activo
    for activo in activos:
        # Configurar fechas (últimos 2 años)
        fecha_inicio = datetime.now() - timedelta(days=730)
        fechas = pd.date_range(start=fecha_inicio, periods=500, freq='D')
        
        # Generar precios realistas usando random walk
        np.random.seed(hash(activo) % 1000)  # Seed basada en el nombre para consistencia
        precio_inicial = np.random.uniform(50, 300)
        retornos = np.random.normal(0.0005, 0.02, len(fechas))  # Retornos diarios
        precios = [precio_inicial]
        
        for retorno in retornos[1:]:
            nuevo_precio = precios[-1] * (1 + retorno)
            precios.append(max(nuevo_precio, 1))  # Evitar precios negativos
        
        # Crear DataFrame
        df = pd.DataFrame({
            'Fecha': fechas,
            'Precio': precios,
            'Volumen': np.random.randint(1000000, 10000000, len(fechas))
        })
        
        # Guardar archivo
        archivo_salida = os.path.join(CARPETA_DATOS, f'{activo}.xlsx')
        df.to_excel(archivo_salida, index=False)
        print(f"   Generado: {archivo_salida}")

# ========================= ANÁLISIS FINANCIERO =========================

def crear_portada_pdf(pdf, tickers):
    """Crea una portada profesional para el PDF"""
    fig = plt.figure(figsize=(8.27, 11.69))  # Tamaño A4
    fig.patch.set_facecolor('white')
    
    # Configurar el layout sin ejes
    ax = fig.add_subplot(111)
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis('off')
    
    # Encabezado principal
    ax.text(0.5, 0.85, 'ANÁLISIS FINANCIERO', 
            fontsize=24, weight='bold', ha='center', va='center',
            color='#1a237e', transform=ax.transAxes)
    
    # Línea decorativa
    ax.plot([0.1, 0.9], [0.82, 0.82], color='#1a237e', linewidth=2, transform=ax.transAxes)
    
    # Subtítulo con activos
    ax.text(0.5, 0.78, f'Activos analizados: {", ".join(tickers)}', 
            fontsize=14, ha='center', va='center', style='italic',
            color='#424242', transform=ax.transAxes)
    
    # Ecuación matemática relevante
    ax.text(0.5, 0.72, 'Fórmula de retorno anualizado:', 
            fontsize=12, ha='center', va='center', weight='bold',
            color='#d32f2f', transform=ax.transAxes)
    
    # Ecuación (sin LaTeX para evitar problemas)
    ax.text(0.5, 0.68, 'R_anual = (1 + R_diario_promedio)^252 - 1', 
            fontsize=14, ha='center', va='center', style='italic',
            color='#d32f2f', transform=ax.transAxes, 
            bbox=dict(boxstyle="round,pad=0.3", facecolor='#ffebee', alpha=0.8))
    
    # Descripción del contenido
    descripcion = """CONTENIDO DEL INFORME:

• Análisis de precios históricos con medias móviles (21d, 63d, 252d)
• Cálculo de retornos diarios y métricas de volatilidad
• Distribución estadística de rendimientos
• Análisis técnico automatizado por activo
• Comparación de rendimientos entre activos
• Recomendaciones técnicas basadas en indicadores

METODOLOGÍA:

• Medias móviles simples para identificación de tendencias
• Análisis de volatilidad para evaluación de riesgo
• Cálculo de retornos anualizados con base en 252 días hábiles
• Análisis estadístico de distribución de retornos"""
    
    ax.text(0.1, 0.58, descripcion, 
            fontsize=11, ha='left', va='top',
            color='#424242', transform=ax.transAxes)
    
    # Información del autor
    ax.text(0.5, 0.25, f'Desarrollado por: {AUTOR_NOMBRE}', 
            fontsize=14, ha='center', va='center', weight='bold',
            color='#1a237e', transform=ax.transAxes)
    
    ax.text(0.5, 0.21, f'{AUTOR_PORTFOLIO}', 
            fontsize=12, ha='center', va='center', style='italic',
            color='#424242', transform=ax.transAxes)
    
    # Fecha de generación
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    ax.text(0.5, 0.15, f'Fecha de análisis: {fecha_actual}', 
            fontsize=11, ha='center', va='center',
            color='#424242', transform=ax.transAxes)
    
    # Espacio reservado - sin nota adicional para mantener diseño limpio
    
    # Pie de página (sin interferir con contenido)
    fig.text(0.5, 0.02, AUTOR_PORTFOLIO, 
            ha='center', va='bottom', fontsize=9, color='#1a237e')
    
    plt.tight_layout()
    pdf.savefig(fig, bbox_inches='tight')
    plt.close(fig)

def agregar_pie_pagina_correcto(fig, numero_pagina, total_paginas):
    """Función deshabilitada - sin pie de página para mejor estética"""
    # No agregamos pie de página para mantener diseño limpio
    pass

def crear_grafico_precios(df, nombre, usar_log=False, numero_pagina=1, total_paginas=1):
    """Crea gráfico de precios SIN superposición con pie de página"""
    # TAMAÑO REDUCIDO para que quepa perfecto en A4
    fig, ax = plt.subplots(figsize=(7.5, 4.5))  
    
    # Configurar escala
    if usar_log:
        ax.set_yscale('log')
        titulo_extra = " (Escala Logarítmica)"
    else:
        titulo_extra = " (Escala Linear)"
    
    # Plotear datos
    df['Precio'].plot(ax=ax, label='Precio', linewidth=1.5, color='#1976d2')
    df['SMA_21'].plot(ax=ax, label='SMA 21d', linewidth=1, color='#d32f2f')
    df['SMA_63'].plot(ax=ax, label='SMA 63d', linewidth=1, color='#f57c00')
    df['SMA_252'].plot(ax=ax, label='SMA 252d', linewidth=1.5, color='#388e3c')
    
    ax.set_title(f'{nombre} - Análisis de Precios{titulo_extra}', 
                fontsize=12, weight='bold', color='#1a237e')
    ax.set_ylabel("Precio (USD)", fontsize=10)
    ax.set_xlabel("Fecha", fontsize=10)
    ax.legend(loc='upper left', fontsize=8)
    ax.grid(True, alpha=0.3)
    
    # CRUCIAL: Dejar espacio para pie de página
    plt.tight_layout()
    plt.subplots_adjust(bottom=0.15)  # Espacio para pie de página
    
    # Agregar pie de página SIN superposición
    agregar_pie_pagina_correcto(fig, numero_pagina, total_paginas)
    
    return fig

def crear_grafico_retornos(df, nombre, numero_pagina=1, total_paginas=1):
    """Crea gráfico de retornos SIN superposición"""
    fig, ax = plt.subplots(figsize=(7.5, 4))
    
    df['Retorno'].plot(ax=ax, color='#ff6f00', alpha=0.7, linewidth=0.8)
    ax.axhline(0, color='#424242', linestyle='--', linewidth=1, alpha=0.8)
    
    # Agregar bandas de volatilidad
    retorno_medio = df['Retorno'].mean()
    vol_std = df['Retorno'].std()
    ax.axhline(retorno_medio + vol_std, color='#d32f2f', linestyle=':', alpha=0.6, label='+1σ')
    ax.axhline(retorno_medio - vol_std, color='#d32f2f', linestyle=':', alpha=0.6, label='-1σ')
    
    ax.set_title(f'{nombre} - Retornos Diarios', fontsize=12, weight='bold', color='#1a237e')
    ax.set_ylabel("Retorno (%)", fontsize=10)
    ax.set_xlabel("Fecha", fontsize=10)
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)
    
    # Espacio para pie de página
    plt.tight_layout()
    plt.subplots_adjust(bottom=0.15)
    
    agregar_pie_pagina_correcto(fig, numero_pagina, total_paginas)
    
    return fig

def crear_histograma_retornos(df, nombre, numero_pagina=1, total_paginas=1):
    """Crea histograma SIN superposición"""
    fig, ax = plt.subplots(figsize=(7.5, 4.5))
    
    retornos_limpios = df['Retorno'].dropna()
    
    # Histograma con curva de densidad
    sns.histplot(retornos_limpios, bins=50, kde=True, ax=ax, 
                color='#00695c', alpha=0.7, stat='density')
    
    # Estadísticas
    mu = retornos_limpios.mean()
    sigma = retornos_limpios.std()
    
    # Líneas de referencia
    ax.axvline(mu, color='#d32f2f', linestyle='--', linewidth=2, label=f'Media: {mu:.4f}')
    ax.axvline(mu + sigma, color='#388e3c', linestyle='--', linewidth=1.5, label=f'+1σ: {mu + sigma:.4f}')
    ax.axvline(mu - sigma, color='#388e3c', linestyle='--', linewidth=1.5, label=f'-1σ: {mu - sigma:.4f}')
    
    ax.set_title(f'{nombre} - Distribución de Retornos', fontsize=12, weight='bold', color='#1a237e')
    ax.set_xlabel("Retorno Diario", fontsize=10)
    ax.set_ylabel("Densidad", fontsize=10)
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)
    
    # Espacio para pie de página
    plt.tight_layout()
    plt.subplots_adjust(bottom=0.15)
    
    agregar_pie_pagina_correcto(fig, numero_pagina, total_paginas)
    
    return fig

def crear_comentario_activo(nombre, stats, df, numero_pagina=1, total_paginas=1):
    """Crea página de comentarios para cada activo"""
    fig, ax = plt.subplots(figsize=(8.27, 10.5))  # Reducido para dejar espacio al pie
    ax.axis('off')
    
    # Título del activo
    ax.text(0.5, 0.95, f'ANÁLISIS TÉCNICO: {nombre}', 
            fontsize=16, weight='bold', ha='center', va='top',
            color='#1a237e', transform=ax.transAxes)
    
    # Línea decorativa
    ax.plot([0.1, 0.9], [0.92, 0.92], color='#1a237e', linewidth=1.5, transform=ax.transAxes)
    
    # Métricas principales
    metricas_texto = f"""MÉTRICAS FINANCIERAS PRINCIPALES:

Precio Actual: ${stats['precio_actual']:.2f} USD
Retorno Anual Estimado: {stats['retorno_anual']*100:.2f}%
Volatilidad Diaria: {stats['volatilidad_diaria']*100:.2f}%
Volatilidad Mensual: {stats['volatilidad_mensual']*100:.2f}%

MEDIAS MÓVILES:
• SMA 21 días: ${stats['sma_21']:.2f}
• SMA 63 días: ${stats['sma_63']:.2f} 
• SMA 252 días: ${stats['sma_252']:.2f}

ANÁLISIS TÉCNICO:"""
    
    # Análisis de tendencia
    precio_actual = stats['precio_actual']
    sma_252 = stats['sma_252']
    sma_63 = stats['sma_63']
    sma_21 = stats['sma_21']
    
    if precio_actual > sma_252:
        tendencia = "ALCISTA - El precio está por encima de la media móvil anual"
        tendencia_simbolo = "↗"
    else:
        tendencia = "BAJISTA - El precio está por debajo de la media móvil anual"
        tendencia_simbolo = "↘"
    
    if sma_21 > sma_63 > sma_252:
        momentum = "MOMENTUM POSITIVO - Todas las medias en orden ascendente"
        momentum_simbolo = "↑"
    elif sma_21 < sma_63 < sma_252:
        momentum = "MOMENTUM NEGATIVO - Todas las medias en orden descendente"
        momentum_simbolo = "↓"
    else:
        momentum = "MOMENTUM MIXTO - Medias móviles entrelazadas"
        momentum_simbolo = "→"
    
    # Análisis de volatilidad
    vol_diaria = stats['volatilidad_diaria']
    if vol_diaria < 0.015:
        vol_nivel = "BAJA"
    elif vol_diaria < 0.025:
        vol_nivel = "MODERADA"
    else:
        vol_nivel = "ALTA"
    
    analisis_texto = f"""
{tendencia_simbolo} Tendencia: {tendencia}

{momentum_simbolo} Momentum: {momentum}

• Volatilidad: {vol_nivel} ({vol_diaria*100:.2f}% diaria)

INTERPRETACIÓN TÉCNICA:

La evaluación técnica de {nombre} se basa en el análisis de medias móviles 
y métricas de volatilidad. El precio actual de ${precio_actual:.2f} en relación 
con las medias móviles indica la dirección predominante del activo.

RECOMENDACIONES TÉCNICAS:

• Soporte técnico estimado: ${min(sma_63, sma_252):.2f}
• Resistencia técnica estimada: ${max(sma_21, precio_actual * 1.05):.2f}
• Nivel de riesgo: {vol_nivel}

CONSIDERACIONES:

El análisis se basa en datos históricos y metodologías de análisis técnico 
reconocidas. Las condiciones del mercado pueden cambiar rápidamente y 
requieren monitoreo constante."""
    
    # Mostrar texto
    ax.text(0.05, 0.88, metricas_texto, 
            fontsize=10, ha='left', va='top',
            color='#212121', transform=ax.transAxes,
            family='monospace')
    
    ax.text(0.05, 0.55, analisis_texto, 
            fontsize=10, ha='left', va='top',
            color='#424242', transform=ax.transAxes)
    
    # Espacio para pie de página
    plt.tight_layout()
    plt.subplots_adjust(bottom=0.08)
    
    agregar_pie_pagina_correcto(fig, numero_pagina, total_paginas)
    
    return fig

def procesar_activos():
    """Procesa todos los activos y genera análisis completo"""
    configurar_matplotlib()
    
    # Verificar si existen datos, si no, generarlos
    if not os.path.exists(CARPETA_DATOS) or not os.listdir(CARPETA_DATOS):
        generar_datos_prueba()
    
    # Crear directorio temporal para gráficos
    os.makedirs(CARPETA_GRAFICOS, exist_ok=True)
    
    # Obtener archivos de datos
    archivos = [f for f in os.listdir(CARPETA_DATOS) if f.endswith('.xlsx') or f.endswith('.xls')]
    
    if not archivos:
        print("No se encontraron archivos de datos. Generando datos de prueba...")
        generar_datos_prueba()
        archivos = [f for f in os.listdir(CARPETA_DATOS) if f.endswith('.xlsx') or f.endswith('.xls')]
    
    tickers = [f.replace('.xlsx', '').replace('.xls', '') for f in archivos]
    print(f"Procesando {len(tickers)} activos: {', '.join(tickers)}")
    
    # Calcular total de páginas aproximado (portada + 4 páginas por activo)
    total_paginas = 1 + (len(tickers) * 4)
    pagina_actual = 1
    
    # Diccionarios para almacenar resultados
    resultados = {}
    dataframes = {}
    
    # Crear PDF
    pdf = PdfPages(ARCHIVO_PDF)
    
    # Crear portada
    print("Creando portada del PDF...")
    crear_portada_pdf(pdf, tickers)
    pagina_actual += 1
    
    # Procesar cada activo
    for archivo in archivos:
        nombre = archivo.replace('.xlsx', '').replace('.xls', '')
        ruta = os.path.join(CARPETA_DATOS, archivo)
        
        print(f"Procesando {nombre}...")
        
        try:
            df = pd.read_excel(ruta)
            
            # Procesar fechas
            if 'Fecha' in df.columns:
                df['Fecha'] = pd.to_datetime(df['Fecha'])
                df.set_index('Fecha', inplace=True)
            elif 'Date' in df.columns:
                df['Date'] = pd.to_datetime(df['Date'])
                df.set_index('Date', inplace=True)
            else:
                print(f"  {nombre}: No se encontró columna de fecha")
                continue
            
            # Procesar precios
            if 'Precio_Cierre' in df.columns:
                df.rename(columns={'Precio_Cierre': 'Precio'}, inplace=True)
            elif 'Adj Close' in df.columns:
                df.rename(columns={'Adj Close': 'Precio'}, inplace=True)
            elif 'Close' in df.columns:
                df.rename(columns={'Close': 'Precio'}, inplace=True)
            elif 'Precio' in df.columns:
                pass  # Ya tiene el nombre correcto
            else:
                print(f"  {nombre}: No se encontró columna de precio")
                continue
            
            # Calcular indicadores técnicos
            df['Retorno'] = df['Precio'].pct_change()
            df['RetornoSemanal'] = df['Retorno'].rolling(window=5).mean()
            df['SMA_21'] = df['Precio'].rolling(window=21).mean()
            df['SMA_63'] = df['Precio'].rolling(window=63).mean()
            df['SMA_252'] = df['Precio'].rolling(window=252).mean()
            
            # Calcular estadísticas
            retorno_diario = df['Retorno'].mean()
            retorno_semanal = df['Retorno'].rolling(5).mean().mean()
            retorno_mensual = df['Retorno'].rolling(21).mean().mean()
            retorno_anual = (1 + retorno_diario) ** 252 - 1
            
            vol_diaria = df['Retorno'].std()
            vol_semanal = df['Retorno'].rolling(5).std().mean()
            vol_mensual = df['Retorno'].rolling(21).std().mean()
            
            stats = {
                'retorno_diario_prom': retorno_diario,
                'retorno_semanal_prom': retorno_semanal,
                'retorno_mensual_prom': retorno_mensual,
                'retorno_anual': retorno_anual,
                'volatilidad_diaria': vol_diaria,
                'volatilidad_semanal': vol_semanal,
                'volatilidad_mensual': vol_mensual,
                'precio_actual': df['Precio'].iloc[-1],
                'sma_21': df['SMA_21'].iloc[-1],
                'sma_63': df['SMA_63'].iloc[-1],
                'sma_252': df['SMA_252'].iloc[-1]
            }
            
            resultados[nombre] = stats
            dataframes[nombre] = df
            
            # Crear gráficos
            print(f"   Creando gráficos para {nombre}...")
            
            # Gráfico de precios
            fig_precio = crear_grafico_precios(df, nombre, USAR_ESCALA_LOGARITMICA, pagina_actual, total_paginas)
            fig_precio.savefig(f'{CARPETA_GRAFICOS}/{nombre}_precio.png', dpi=300, bbox_inches='tight')
            pdf.savefig(fig_precio, bbox_inches='tight', pad_inches=0.1)
            plt.close(fig_precio)
            pagina_actual += 1
            
            # Comentario del activo
            fig_comentario = crear_comentario_activo(nombre, stats, df, pagina_actual, total_paginas)
            pdf.savefig(fig_comentario, bbox_inches='tight', pad_inches=0.1)
            plt.close(fig_comentario)
            pagina_actual += 1
            
            # Gráfico de retornos
            fig_retorno = crear_grafico_retornos(df, nombre, pagina_actual, total_paginas)
            fig_retorno.savefig(f'{CARPETA_GRAFICOS}/{nombre}_retorno.png', dpi=300, bbox_inches='tight')
            pdf.savefig(fig_retorno, bbox_inches='tight', pad_inches=0.1)
            plt.close(fig_retorno)
            pagina_actual += 1
            
            # Histograma
            fig_hist = crear_histograma_retornos(df, nombre, pagina_actual, total_paginas)
            fig_hist.savefig(f'{CARPETA_GRAFICOS}/{nombre}_histograma.png', dpi=300, bbox_inches='tight')
            pdf.savefig(fig_hist, bbox_inches='tight', pad_inches=0.1)
            plt.close(fig_hist)
            pagina_actual += 1
            
            print(f"   {nombre} procesado correctamente")
            
        except Exception as e:
            print(f"Error procesando {nombre}: {str(e)}")
            continue
    
    pdf.close()
    print(f"PDF guardado como: {ARCHIVO_PDF}")
    
    return resultados, dataframes

def crear_excel_REALMENTE_corregido(resultados, dataframes):
    """Crea archivo Excel con VERDADERA separación entre imágenes y datos"""
    print("Creando archivo Excel con VERDADERA separación...")
    
    # Crear DataFrame de resumen
    resumen = pd.DataFrame(resultados).T
    
    # Crear gráfico comparativo primero
    fig, ax = plt.subplots(figsize=(12, 8))
    retornos_sorted = resumen['retorno_anual'].sort_values()
    colors = ['#d32f2f' if x < 0 else '#388e3c' for x in retornos_sorted]
    
    bars = retornos_sorted.plot(kind='barh', ax=ax, color=colors, figsize=(12, 8))
    ax.set_title('Comparación de Retornos Anuales Estimados', fontsize=16, weight='bold', color='#1a237e')
    ax.set_xlabel('Retorno Anual Estimado (%)', fontsize=12)
    ax.grid(True, alpha=0.3, axis='x')
    
    # Agregar valores en las barras
    for i, v in enumerate(retornos_sorted):
        ax.text(v + (0.005 if v >= 0 else -0.005), i, f'{v*100:.1f}%', 
               va='center', ha='left' if v >= 0 else 'right', fontsize=10, weight='bold')
    
    # Agregar información del autor
    ax.text(0.02, 0.98, f'Desarrollado por: {AUTOR_NOMBRE} | {AUTOR_PORTFOLIO}', 
            transform=ax.transAxes, fontsize=10, va='top', style='italic', color='#666666')
    
    plt.tight_layout()
    plt.savefig(f'{CARPETA_GRAFICOS}/resumen_retornos.png', dpi=300, bbox_inches='tight')
    plt.close(fig)
    
    # Crear Excel
    with pd.ExcelWriter(ARCHIVO_EXCEL, engine='openpyxl') as writer:
        # HOJA RESUMEN: Tabla MUCHO MÁS ABAJO para evitar superposición
        resumen.to_excel(writer, sheet_name='Resumen', startrow=35)  # Fila 35 en lugar de 25
        
        # HOJAS INDIVIDUALES: Datos MUCHO MÁS ABAJO
        for nombre, df in dataframes.items():
            print(f"   Creando hoja para {nombre}...")
            # Datos empiezan en fila 80 (antes era 60) para evitar superposición
            df.to_excel(writer, sheet_name=nombre, startrow=80)
    
    # Cargar el workbook para agregar imágenes
    workbook = load_workbook(ARCHIVO_EXCEL)
    
    # HOJA DE RESUMEN: imagen arriba, tabla MUCHO más abajo
    ws_resumen = workbook['Resumen']
    
    # Títulos de la hoja de resumen
    ws_resumen['A1'] = 'ANÁLISIS FINANCIERO'
    ws_resumen['A1'].font = Font(size=16, bold=True, color='1a237e')
    ws_resumen['A2'] = f'Desarrollado por: {AUTOR_NOMBRE}'
    ws_resumen['A2'].font = Font(size=12, italic=True)
    ws_resumen['A3'] = f'Portfolio: {AUTOR_PORTFOLIO}'
    ws_resumen['A3'].font = Font(size=12, italic=True, color='666666')
    
    # Imagen de resumen (arriba, tabla en fila 35)
    img_resumen = Image(f'{CARPETA_GRAFICOS}/resumen_retornos.png')
    img_resumen.width = 900
    img_resumen.height = 600
    ws_resumen.add_image(img_resumen, 'A5')  # Imagen en A5, tabla en fila 35
    
    # HOJAS INDIVIDUALES: Gráficos arriba, datos en fila 80
    for nombre, df in dataframes.items():
        ws = workbook[nombre]
        
        # Título de la hoja
        ws['A1'] = f'ANÁLISIS DE {nombre}'
        ws['A1'].font = Font(size=14, bold=True, color='1a237e')
        ws['A2'] = f'{AUTOR_PORTFOLIO}'
        ws['A2'].font = Font(size=10, italic=True, color='666666')
        
        # Separación clara: gráficos de la fila 4 a 75, datos desde fila 80
        row_inicio_graficos = 4
        
        # Gráfico de precios
        if os.path.exists(f'{CARPETA_GRAFICOS}/{nombre}_precio.png'):
            img_precio = Image(f'{CARPETA_GRAFICOS}/{nombre}_precio.png')
            img_precio.width = 700
            img_precio.height = 300
            ws.add_image(img_precio, f'A{row_inicio_graficos}')
        
        # Gráfico de retornos (debajo del anterior)
        if os.path.exists(f'{CARPETA_GRAFICOS}/{nombre}_retorno.png'):
            img_retorno = Image(f'{CARPETA_GRAFICOS}/{nombre}_retorno.png')
            img_retorno.width = 700
            img_retorno.height = 250
            ws.add_image(img_retorno, f'A{row_inicio_graficos + 20}')
        
        # Histograma (más abajo)
        if os.path.exists(f'{CARPETA_GRAFICOS}/{nombre}_histograma.png'):
            img_hist = Image(f'{CARPETA_GRAFICOS}/{nombre}_histograma.png')
            img_hist.width = 600
            img_hist.height = 300
            ws.add_image(img_hist, f'A{row_inicio_graficos + 40}')
    
    workbook.save(ARCHIVO_EXCEL)
    print(f"Excel guardado como: {ARCHIVO_EXCEL}")

def main():
    """Función principal"""
    print("="*60)
    print("ANÁLISIS FINANCIERO FINAL - TODOS LOS PROBLEMAS CORREGIDOS")
    print(f"Desarrollado por: {AUTOR_NOMBRE}")
    print(f"Portfolio: {AUTOR_PORTFOLIO}")
    print("="*60)
    print(f"Configuración:")
    print(f"   • Escala de gráficos: {'Logarítmica' if USAR_ESCALA_LOGARITMICA else 'Linear'}")
    print(f"   • Carpeta de datos: {CARPETA_DATOS}")
    print(f"   • Archivo PDF: {ARCHIVO_PDF}")
    print(f"   • Archivo Excel: {ARCHIVO_EXCEL}")
    print("="*60)
    
    try:
        # Procesar activos
        resultados, dataframes = procesar_activos()
        
        if not resultados:
            print("No se pudieron procesar activos")
            return
        
        # Crear Excel REALMENTE corregido
        crear_excel_REALMENTE_corregido(resultados, dataframes)
        
        # Limpiar archivos temporales
        if os.path.exists(CARPETA_GRAFICOS):
            shutil.rmtree(CARPETA_GRAFICOS)
        
        print("\n" + "="*60)
        print("ANÁLISIS FINAL COMPLETADO - TODOS LOS PROBLEMAS SOLUCIONADOS")
        print("="*60)
        print(f"Archivos generados:")
        print(f"   • {ARCHIVO_PDF} (A4 perfecto, SIN superposición de pie de página)")
        print(f"   • {ARCHIVO_EXCEL} (imágenes REALMENTE separadas de datos)")
        print(f"Activos procesados: {len(resultados)}")
        print(f"Desarrollado por: {AUTOR_NOMBRE} | {AUTOR_PORTFOLIO}")
        print("="*60)
        
    except Exception as e:
        print(f"Error en el análisis: {str(e)}")
        raise

if __name__ == "__main__":
    main()