import os
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from fpdf import FPDF
from datetime import datetime
from textwrap import wrap

# ------------------------ CONFIGURACIÓN ------------------------
CARPETA_DATOS = './datospython1'
CARPETA_SALIDA = './graficos_temp'
os.makedirs(CARPETA_SALIDA, exist_ok=True)

# ------------------------ LECTURA Y PREPARACIÓN ------------------------
dataframes = {}
for archivo in os.listdir(CARPETA_DATOS):
    ruta = os.path.join(CARPETA_DATOS, archivo)
    nombre = os.path.splitext(archivo)[0]

    try:
        if archivo.endswith('.csv'):
            df = pd.read_csv(ruta, index_col=0, parse_dates=True)
        elif archivo.endswith('.xlsx') or archivo.endswith('.xls'):
            df = pd.read_excel(ruta, index_col=0, parse_dates=True)
        else:
            continue

        columnas = [col for col in df.columns if col.lower() in ['adj close', 'close', 'precio_cierre']]
        if not columnas:
            continue

        df_filtrado = df[[columnas[0]]].rename(columns={columnas[0]: nombre})
        dataframes[nombre] = df_filtrado

    except Exception as e:
        print(f"⚠️ Error con {archivo}: {e}")

if not dataframes:
    print("❌ No se encontraron datos válidos.")
    exit()

# ------------------------ MATRIZ DE CORRELACIÓN ------------------------
df_precios = pd.concat(dataframes.values(), axis=1).dropna()

BENCHMARK = 'SPY'
if BENCHMARK in df_precios.columns:
    df_precios = df_precios.drop(columns=BENCHMARK)

retornos = np.log(df_precios / df_precios.shift(1)).dropna()
matriz_correlacion = retornos.corr()

# ------------------------ EXCEL ------------------------
excel_path = os.path.join(CARPETA_SALIDA, 'matriz_correlacion.xlsx')
wb = Workbook()
ws = wb.active
ws.title = "Correlaciones"
for fila in dataframe_to_rows(matriz_correlacion.round(4), index=True, header=True):
    ws.append(fila)
wb.save(excel_path)

# ------------------------ HEATMAP ------------------------
plt.figure(figsize=(12, 10))
sns.set(style="whitegrid", font_scale=1.1)
sns.heatmap(matriz_correlacion, annot=True, fmt=".2f", cmap="coolwarm", vmin=-1, vmax=1,
            square=True, linewidths=0.5, cbar_kws={'label': 'Correlación'})
plt.title("Matriz de Correlación entre Activos", fontsize=16, weight='bold')
plt.xticks(rotation=45, ha='right')
plt.yticks(rotation=0)
plt.tight_layout()
img_path = os.path.join(CARPETA_SALIDA, 'heatmap_correlacion.png')
plt.savefig(img_path)
plt.close()

print(f"\n✅ Matriz de correlación generada correctamente:")
print(f"- 📄 Excel: {excel_path}")
print(f"- 🗾️ Imagen: {img_path}")
print("\n🔎 Vista previa:")
print(matriz_correlacion.round(2))

# ------------------------ INSIGHTS AUTOMÁTICOS ------------------------
insights = []
for i in range(len(matriz_correlacion.columns)):
    for j in range(i+1, len(matriz_correlacion.columns)):
        activo1 = matriz_correlacion.columns[i]
        activo2 = matriz_correlacion.columns[j]
        correlacion = matriz_correlacion.iloc[i, j]

        if correlacion >= 0.8:
            insights.append(f"Alta correlación positiva entre {activo1} y {activo2} ({correlacion:.2f}). Se mueven casi en sincronía.")
        elif correlacion <= -0.5:
            insights.append(f"Correlación negativa entre {activo1} y {activo2} ({correlacion:.2f}). Tienden a moverse en direcciones opuestas.")
        elif -0.2 <= correlacion <= 0.2:
            insights.append(f"Baja correlación entre {activo1} y {activo2} ({correlacion:.2f}). Buena combinación para diversificación.")

print("\n📊 Insights automáticos:")
for linea in insights:
    print(" - " + linea)

insight_file = os.path.join(CARPETA_SALIDA, 'insights_correlacion.txt')
with open(insight_file, 'w', encoding='utf-8') as f:
    f.write("Insights Cuantitativos sobre Correlaciones\n")
    f.write("="*50 + "\n\n")
    for linea in insights:
        f.write(linea + "\n")

print(f"\n🗘️ Resumen guardado en: {insight_file}")

# ------------------------ CLASIFICACIÓN POR SECTOR ------------------------
sectores_path = os.path.join(CARPETA_DATOS, 'sectores.csv')
tabla_sector = ""
if os.path.exists(sectores_path):
    df_sectores = pd.read_csv(sectores_path)
    df_sectores.columns = df_sectores.columns.str.strip()

    posibles_col_activo = ['Activo', 'Ticker', 'Symbol']
    posibles_col_sector = ['Sector', 'GICS Sector']

    col_activo = next((col for col in posibles_col_activo if col in df_sectores.columns), None)
    col_sector = next((col for col in posibles_col_sector if col in df_sectores.columns), None)

    if col_activo and col_sector:
        mapa_sectores = dict(zip(df_sectores[col_activo], df_sectores[col_sector]))

        activos_por_sector = {}
        for activo in matriz_correlacion.columns:
            sector = mapa_sectores.get(activo, 'Sin clasificar')
            activos_por_sector.setdefault(sector, []).append(activo)

        sector_correlations = []
        for sector, activos in activos_por_sector.items():
            if len(activos) < 2:
                continue
            submatriz = matriz_correlacion.loc[activos, activos]
            cor_media = submatriz.where(~np.eye(len(submatriz), dtype=bool)).mean().mean()
            sector_correlations.append((sector, round(cor_media, 3)))

        print("\n🏷️ Promedio de correlación intra-sectorial:")
        tabla_sector += "\nPromedio de correlación intra-sectorial:\n"
        for sector, media in sector_correlations:
            print(f" - {sector}: {media}")
            tabla_sector += f" - {sector}: {media}\n"
    else:
        print("\n⚠️ El archivo de sectores no contiene las columnas necesarias para clasificar.")
else:
    print("\n📂 No se encontró el archivo 'sectores.csv'.")
    print("ℹ️ Si lo agregás en ./datospython1, podré calcular correlaciones por sector y sugerencias de cobertura sectorial.")
    print("Ejemplo de columnas reconocidas: Ticker / Symbol y Sector / GICS Sector")

# ------------------------ PDF PROFESIONAL ------------------------
pdf_path = os.path.join(CARPETA_SALIDA, "informe_correlacion_financiera.pdf")
pdf = FPDF()
pdf.set_auto_page_break(auto=True, margin=15)
pdf.add_page()

pdf.set_font("helvetica", 'B', 16)
pdf.cell(0, 10, "Informe de Correlación Financiera", ln=True, align="C")
pdf.set_font("helvetica", '', 12)
pdf.cell(0, 10, "Autor: Leonardo Caliva", ln=True, align="C")
pdf.cell(0, 10, f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y')}", ln=True, align="C")
pdf.cell(0, 10, "Sitio web: https://leocaliva.com", ln=True, align="C")
pdf.ln(10)

pdf.set_font("helvetica", 'B', 13)
pdf.cell(0, 10, "Resumen Ejecutivo", ln=True)
pdf.set_font("helvetica", '', 11)
resumen = (
    "Este informe presenta las correlaciones históricas entre activos financieros seleccionados, "
    "permitiendo identificar relaciones de dependencia y oportunidades de diversificación dentro de una cartera. "
    f"El benchmark utilizado fue {BENCHMARK}, el cual fue excluido del análisis de correlación entre activos para evitar distorsiones."
)
pdf.multi_cell(0, 8, resumen)
pdf.ln(5)

if os.path.exists(img_path):
    pdf.set_font("helvetica", 'B', 14)
    pdf.cell(0, 10, "Matriz de Correlación entre Activos", ln=True)
    pdf.image(img_path, x=15, w=180)
    pdf.ln(10)

pdf.add_page()
pdf.set_font("helvetica", 'B', 14)
pdf.cell(0, 10, "Insights Cuantitativos sobre Correlaciones", ln=True)
pdf.set_font("helvetica", '', 11)
pdf.cell(0, 8, "="*60, ln=True)
pdf.ln(2)

for line in insights:
    pdf.multi_cell(0, 8, txt=line)

if tabla_sector:
    pdf.add_page()
    pdf.set_font("helvetica", 'B', 14)
    pdf.cell(0, 10, "Resumen de Correlación por Sector", ln=True)
    pdf.set_font("helvetica", '', 11)
    for linea in tabla_sector.splitlines():
        pdf.multi_cell(0, 8, linea)

try:
    pdf.output(pdf_path)
    print(f"\n📄 PDF profesional generado en: {pdf_path}")
except Exception as e:
    print(f"❌ Error al generar el PDF: {e}")
