import os
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from fpdf import FPDF
from datetime import datetime
import scipy.cluster.hierarchy as sch
from scipy.cluster.hierarchy import linkage, dendrogram, fcluster
from sklearn.decomposition import PCA
import networkx as nx

# ------------------------ CONFIGURACIÓN ------------------------
CARPETA_DATOS = './datospython1'
CARPETA_SALIDA = './graficos_temp'
BENCHMARK = 'SPY' # Define el ticker del benchmark.
os.makedirs(CARPETA_SALIDA, exist_ok=True)

# Helper para codificar texto para FPDF, crucial para evitar errores
def encode_text(text):
    return str(text).encode('latin-1', 'replace').decode('latin-1')

# ------------------------ LECTURA Y PREPARACIÓN DE DATOS ------------------------
print(f"Directorio actual: {os.getcwd()}")
print(f"Buscando datos en: {os.path.abspath(CARPETA_DATOS)}")

dataframes = {}
for archivo in os.listdir(CARPETA_DATOS):
    ruta = os.path.join(CARPETA_DATOS, archivo)
    nombre = os.path.splitext(archivo)[0]
    try:
        if archivo.endswith('.csv'):
            df = pd.read_csv(ruta, index_col=0, parse_dates=True)
        elif archivo.endswith('.xlsx') or archivo.endswith('.xls'):
            df = pd.read_excel(ruta, index_col=0, parse_dates=True)
        else:
            continue
        columnas = [col for col in df.columns if col.lower() in ['adj close', 'close', 'precio_cierre']]
        if not columnas:
            continue
        df_filtrado = df[[columnas[0]]].rename(columns={columnas[0]: nombre})
        dataframes[nombre] = df_filtrado
    except Exception as e:
        print(f"⚠️ Error con {archivo}: {e}")

if not dataframes:
    print("❌ No se encontraron datos válidos.")
    exit()

df_precios_full = pd.concat(dataframes.values(), axis=1).dropna()
retornos_full = np.log(df_precios_full / df_precios_full.shift(1)).dropna()

# Separar benchmark del resto de los activos
if BENCHMARK in retornos_full.columns:
    retornos_benchmark = retornos_full[[BENCHMARK]]
    retornos_activos = retornos_full.drop(columns=BENCHMARK)
    print(f"✅ Benchmark '{BENCHMARK}' separado correctamente.")
else:
    retornos_benchmark = None
    retornos_activos = retornos_full
    print(f"⚠️ No se encontró el benchmark '{BENCHMARK}'.")

# La matriz de correlación principal se calcula sobre los activos (sin el benchmark)
matriz_correlacion = retornos_activos.corr()


# =================================================================================
# BLOQUE DE NUEVOS ANÁLISIS SOLICITADOS (Se insertan aquí)
# =================================================================================

# NUEVO 1: ANÁLISIS DE RUPTURAS DE CORRELACIÓN ROLLING
print("\n🔎 Analizando rupturas de correlación recientes...")
rupturas_recientes_texto = []
try:
    ventana_actual_30d = retornos_activos.last('30D')
    fecha_fin_pasada_30d = ventana_actual_30d.index.min() - pd.Timedelta(days=180)
    fecha_inicio_pasada_30d = fecha_fin_pasada_30d - pd.Timedelta(days=30)
    ventana_pasada_30d = retornos_activos.loc[fecha_inicio_pasada_30d:fecha_fin_pasada_30d]

    if not ventana_actual_30d.empty and not ventana_pasada_30d.empty:
        corr_actual_30d = ventana_actual_30d.corr()
        corr_pasada_30d = ventana_pasada_30d.corr()
        deltas = []
        for i in range(len(corr_actual_30d.columns)):
            for j in range(i + 1, len(corr_actual_30d.columns)):
                a1 = corr_actual_30d.columns[i]; a2 = corr_actual_30d.columns[j]
                if a1 in corr_pasada_30d.columns and a2 in corr_pasada_30d.columns:
                    actual = corr_actual_30d.loc[a1, a2]; pasado = corr_pasada_30d.loc[a1, a2]
                    delta = actual - pasado
                    deltas.append(((a1, a2), pasado, actual, delta))
        
        top_5_rupturas = sorted(deltas, key=lambda item: abs(item[3]), reverse=True)[:5]
        if top_5_rupturas:
            for (a1, a2), pasado, actual, delta in top_5_rupturas:
                cambio = "aumento" if delta > 0 else "disminuyo"
                linea = f"La correlacion entre {a1} y {a2} {cambio} de {pasado:.2f} a {actual:.2f} (Delta = {delta:+.2f})."
                rupturas_recientes_texto.append(linea)
    else:
        rupturas_recientes_texto.append("No hay suficientes datos historicos para calcular rupturas de correlacion.")
except Exception as e:
    rupturas_recientes_texto.append(f"Error al calcular rupturas: {e}")

# NUEVO 2: ANÁLISIS DE CLUSTERS (TEXTUAL)
print("\n clustering de activos para resumen textual...")
cluster_summary_texto = []
try:
    linkage_matrix_cluster = sch.linkage(1 - matriz_correlacion.abs(), method='ward')
    NUM_CLUSTERS = 4 
    clusters = fcluster(linkage_matrix_cluster, t=NUM_CLUSTERS, criterion='maxclust')
    activos_en_cluster = pd.DataFrame({'activo': matriz_correlacion.columns, 'cluster_id': clusters})
    
    cluster_summary_texto.append(f"Se detectaron {activos_en_cluster['cluster_id'].nunique()} clusters principales:")
    for i in sorted(activos_en_cluster['cluster_id'].unique()):
        activos = activos_en_cluster[activos_en_cluster['cluster_id'] == i]['activo'].tolist()
        linea = f" - Cluster {i}: Incluye activos como {', '.join(activos[:5])}"
        if len(activos) > 5: linea += " y otros."
        cluster_summary_texto.append(linea)
except Exception as e:
    cluster_summary_texto.append(f"No fue posible realizar el analisis de clusters: {e}")

# NUEVO 3: CORRELACIÓN CONTRA BENCHMARK
print(f"\n📊 Calculando correlación contra el benchmark ({BENCHMARK})...")
corr_benchmark_path = None
corr_benchmark_texto = []
if retornos_benchmark is not None:
    try:
        corr_con_benchmark = retornos_activos.corrwith(retornos_benchmark[BENCHMARK]).sort_values(ascending=False)
        plt.figure(figsize=(10, 8))
        sns.barplot(x=corr_con_benchmark.values, y=corr_con_benchmark.index, orient='h', palette='viridis')
        plt.title(f'Correlacion de Activos contra {BENCHMARK}')
        plt.xlabel('Correlacion'); plt.ylabel('Activo')
        plt.tight_layout()
        corr_benchmark_path = os.path.join(CARPETA_SALIDA, 'corr_vs_benchmark.png')
        plt.savefig(corr_benchmark_path); plt.close()
        top_sensibles = corr_con_benchmark.head(2)
        corr_benchmark_texto.append(f"Los activos mas sensibles al mercado son {top_sensibles.index[0]} (correlacion {top_sensibles.iloc[0]:.2f}) y {top_sensibles.index[1]} ({top_sensibles.iloc[1]:.2f}).")
    except Exception as e:
        print(f"❌ Error en análisis contra benchmark: {e}")

# NUEVO 4: COMPARACIÓN HISTÓRICA (6M vs 6M)
print("\n🔄 Comparando correlaciones: ultimos 6 meses vs 6 meses previos...")
corr_diff_path = None
corr_diff_texto = []
try:
    ventana_actual_6m = retornos_activos.last('180D')
    fecha_inicio_pasada_6m = ventana_actual_6m.index.min() - pd.Timedelta(days=180)
    fecha_fin_pasada_6m = ventana_actual_6m.index.min() - pd.Timedelta(days=1)
    ventana_pasada_6m = retornos_activos.loc[fecha_inicio_pasada_6m:fecha_fin_pasada_6m]

    if not ventana_actual_6m.empty and not ventana_pasada_6m.empty:
        corr_actual_6m = ventana_actual_6m.corr()
        corr_pasada_6m = ventana_pasada_6m.corr()
        corr_pasada_6m, corr_actual_6m = corr_pasada_6m.align(corr_actual_6m, join='inner', axis=1)
        corr_pasada_6m, corr_actual_6m = corr_pasada_6m.align(corr_actual_6m, join='inner', axis=0)
        corr_diff = corr_actual_6m - corr_pasada_6m
        
        plt.figure(figsize=(12, 10))
        sns.heatmap(corr_diff, cmap="RdBu", center=0, annot=False, vmin=-1, vmax=1)
        plt.title('Diferencia de Correlacion (Ultimos 6M vs Anteriores 6M)')
        plt.tight_layout()
        corr_diff_path = os.path.join(CARPETA_SALIDA, 'heatmap_diferencia_6m.png')
        plt.savefig(corr_diff_path); plt.close()
        max_change = corr_diff.stack().abs().idxmax()
        max_delta = corr_diff.loc[max_change]
        corr_diff_texto.append(f"El cambio mas notable fue entre {max_change[0]} y {max_change[1]}, con un delta de {max_delta:.2f}.")
except Exception as e:
    corr_diff_texto.append(f"Error al generar la comparativa historica: {e}")


# =================================================================================
# BLOQUE DE ANÁLISIS Y GRÁFICOS ORIGINALES (Se mantiene intacto)
# =================================================================================
print("\n🎨 Generando visualizaciones originales...")

# ==== Excel ====
excel_path = os.path.join(CARPETA_SALIDA, 'matriz_correlacion.xlsx')
wb = Workbook()
ws = wb.active
ws.title = "Correlaciones"
for fila in dataframe_to_rows(matriz_correlacion.round(4), index=True, header=True):
    ws.append(fila)
wb.save(excel_path)

# ==== Gráfico 1: Matriz de Correlación (Heatmap) ====
linkage_heatmap = sch.linkage(matriz_correlacion, method='ward')
orden = sch.dendrogram(linkage_heatmap, no_plot=True)['leaves']
matriz_correlacion_ordenada = matriz_correlacion.iloc[orden, orden]
img_path = os.path.join(CARPETA_SALIDA, 'heatmap_correlacion.png')
plt.figure(figsize=(12, 10))
sns.heatmap(matriz_correlacion_ordenada, annot=True, fmt=".2f", cmap="coolwarm", vmin=-1, vmax=1, square=True, linewidths=0.5, cbar_kws={'label': 'Correlación'})
plt.title("Matriz de Correlación entre Activos", fontsize=16, weight='bold')
plt.xticks(rotation=45, ha='right'); plt.yticks(rotation=0); plt.tight_layout()
plt.savefig(img_path); plt.close()
print(f"✅ Imagen generada: {img_path}")

# ==== Gráfico 2: Dendrograma ====
dendro_path = os.path.join(CARPETA_SALIDA, 'dendrograma_clustering.png')
linkage_matrix_dendro = linkage(1 - matriz_correlacion.abs(), method='ward')
plt.figure(figsize=(12, 5)); dendrogram(linkage_matrix_dendro, labels=matriz_correlacion.columns, leaf_rotation=45)
plt.title('Dendrograma de Clustering de Activos'); plt.tight_layout()
plt.savefig(dendro_path); plt.close()
print(f"✅ Imagen generada: {dendro_path}")

# ==== Gráfico 3: Network Graph ====
net_path = os.path.join(CARPETA_SALIDA, 'network_graph.png')
plt.figure(figsize=(9, 8)); G = nx.Graph()
for activo in matriz_correlacion.columns: G.add_node(activo)
umbral_cor = 0.3
for i, a1 in enumerate(matriz_correlacion.columns):
    for j, a2 in enumerate(matriz_correlacion.columns):
        if i < j:
            cor_val = matriz_correlacion.loc[a1, a2]
            if abs(cor_val) > umbral_cor: G.add_edge(a1, a2, weight=abs(cor_val))
pos = nx.spring_layout(G, seed=42); edges = G.edges()
weights = [G[u][v]['weight']*3 for u,v in edges]
nx.draw_networkx_nodes(G, pos, node_size=800, node_color='lightblue')
nx.draw_networkx_edges(G, pos, edgelist=edges, width=weights, alpha=0.6)
nx.draw_networkx_labels(G, pos, font_size=10)
plt.title(f'Network Graph de Correlaciones (>|{umbral_cor}|)'); plt.axis('off'); plt.tight_layout()
plt.savefig(net_path); plt.close()
print(f"✅ Imagen generada: {net_path}")

# ==== Gráfico 4: Distribución de Correlaciones ====
hist_path = os.path.join(CARPETA_SALIDA, 'correlation_hist.png')
corrs_flat = matriz_correlacion.where(~np.eye(matriz_correlacion.shape[0], dtype=bool)).stack()
plt.figure(figsize=(8, 5)); sns.histplot(corrs_flat, kde=True, bins=20, color='slateblue')
plt.title('Distribución y densidad de correlaciones (Pearson)'); plt.xlabel('Correlación'); plt.ylabel('Frecuencia')
plt.tight_layout(); plt.savefig(hist_path); plt.close()
print(f"✅ Imagen generada: {hist_path}")

# ==== Gráfico 5: Rolling Correlation (top 2 pares) ====
rolling_paths = []
corrs_abs_sorted = matriz_correlacion.where(~np.eye(matriz_correlacion.shape[0], dtype=bool)).abs().unstack().sort_values(ascending=False)
top_pairs = []; visited_pairs = set()
for (a1, a2) in corrs_abs_sorted.index:
    if a1 != a2 and tuple(sorted((a1, a2))) not in visited_pairs:
        top_pairs.append((a1, a2)); visited_pairs.add(tuple(sorted((a1, a2))))
    if len(top_pairs) == 2: break
ROLLING_WINDOW = 60
for a1, a2 in top_pairs:
    roll_path = os.path.join(CARPETA_SALIDA, f'rollingcorr_{a1}_{a2}.png')
    serie = retornos_activos[a1].rolling(ROLLING_WINDOW).corr(retornos_activos[a2])
    plt.figure(figsize=(10, 4)); plt.plot(serie.index, serie, label=f'{a1}-{a2}')
    plt.axhline(0, color='gray', ls='--'); plt.title(f'Correlación Rolling 60 días: {a1}-{a2}')
    plt.ylabel('Correlación'); plt.xlabel('Fecha'); plt.legend(); plt.tight_layout()
    plt.savefig(roll_path); plt.close(); rolling_paths.append((f"{a1}-{a2}", roll_path))
    print(f"✅ Imagen generada: {roll_path}")

# ==== Gráfico 6: PCA Varianza Explicada ====
pca_path = os.path.join(CARPETA_SALIDA, 'pca_varianza.png')
pca = PCA(); pca.fit(retornos_activos.dropna(axis=1, how='any'))
explained_var = pca.explained_variance_ratio_
plt.figure(figsize=(8, 4)); plt.bar(range(1, len(explained_var)+1), explained_var*100)
plt.xlabel('Componente Principal'); plt.ylabel('% de Varianza Explicada'); plt.title('Varianza Explicada por PCA')
plt.tight_layout(); plt.savefig(pca_path); plt.close()
print(f"✅ Imagen generada: {pca_path}")

# ==== INSIGHTS AUTOMÁTICOS ====
insights = []
for i in range(len(matriz_correlacion.columns)):
    for j in range(i+1, len(matriz_correlacion.columns)):
        activo1 = matriz_correlacion.columns[i]
        activo2 = matriz_correlacion.columns[j]
        correlacion = matriz_correlacion.iloc[i, j]
        if correlacion >= 0.8:
            insights.append(f"Alta correlación positiva entre {activo1} y {activo2} ({correlacion:.2f}). Se mueven casi en sincronía.")
        elif correlacion <= -0.5:
            insights.append(f"Correlación negativa entre {activo1} y {activo2} ({correlacion:.2f}). Tienden a moverse en direcciones opuestas.")
        elif -0.2 <= correlacion <= 0.2:
            insights.append(f"Baja correlación entre {activo1} y {activo2} ({correlacion:.2f}). Buena combinación para diversificación.")

insight_file = os.path.join(CARPETA_SALIDA, 'insights_correlacion.txt')
with open(insight_file, 'w', encoding='utf-8') as f:
    f.write("Insights Cuantitativos sobre Correlaciones\n")
    f.write("="*50 + "\n\n")
    for linea in insights:
        f.write(linea + "\n")

# ==== CLASIFICACIÓN POR SECTOR ====
sectores_path = os.path.join(CARPETA_DATOS, 'sectores.csv')
tabla_sector = ""
sector_correlations = []
mapa_sectores = {}
if os.path.exists(sectores_path):
    df_sectores = pd.read_csv(sectores_path)
    df_sectores.columns = df_sectores.columns.str.strip()
    posibles_col_activo = ['Activo', 'Ticker', 'Symbol']
    posibles_col_sector = ['Sector', 'GICS Sector']
    col_activo = next((col for col in posibles_col_activo if col in df_sectores.columns), None)
    col_sector = next((col for col in posibles_col_sector if col in df_sectores.columns), None)
    if col_activo and col_sector:
        mapa_sectores = dict(zip(df_sectores[col_activo], df_sectores[col_sector]))
        activos_por_sector = {}
        for activo in matriz_correlacion.columns:
            sector = mapa_sectores.get(activo, 'Sin clasificar')
            activos_por_sector.setdefault(sector, []).append(activo)
        for sector, activos in activos_por_sector.items():
            if len(activos) < 2:
                continue
            submatriz = matriz_correlacion.loc[activos, activos]
            cor_media = submatriz.where(~np.eye(len(submatriz), dtype=bool)).mean().mean()
            sector_correlations.append((sector, round(cor_media, 3)))
        tabla_sector += "\nPromedio de correlación intra-sectorial:\n"
        for sector, media in sector_correlations:
            tabla_sector += f" - {sector}: {media}\n"
    else:
        tabla_sector = "\n⚠️ El archivo de sectores no contiene las columnas necesarias para clasificar.\n"
else:
    tabla_sector = "\n📂 No se encontró el archivo 'sectores.csv'.\n"

# ==== EXPORTAR SECTORIAL A EXCEL ====
with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
    if sector_correlations:
        df_sector = pd.DataFrame(sector_correlations, columns=['Sector', 'Correlacion Media'])
        df_sector.to_excel(writer, sheet_name='Correlacion_Sector', index=False)
    if mapa_sectores:
        df_map = pd.DataFrame(list(mapa_sectores.items()), columns=['Activo', 'Sector'])
        df_map.to_excel(writer, sheet_name='Activos_Sector', index=False)
    df_insights = pd.DataFrame(insights, columns=['Insight'])
    df_insights.to_excel(writer, sheet_name='Insights', index=False)

# ==== IMAGEN DE TABLA SECTORIAL ====
sector_img_path = None
if sector_correlations:
    df_sector = pd.DataFrame(sector_correlations, columns=['Sector', 'Correlación Media'])
    fig, ax = plt.subplots(figsize=(6, max(2, len(df_sector)*0.5)))
    ax.axis('tight')
    ax.axis('off')
    table = ax.table(cellText=df_sector.values,
                     colLabels=df_sector.columns,
                     loc='center',
                     cellLoc='center')
    plt.title("Promedio de correlación intra-sectorial", fontsize=14, weight='bold')
    sector_img_path = os.path.join(CARPETA_SALIDA, 'sector_correlation.png')
    plt.savefig(sector_img_path, bbox_inches='tight')
    plt.close()
    print(f"✅ Imagen generada: {sector_img_path}")

# ==== IMAGEN DE INSIGHTS AUTOMÁTICOS ====
insights_img_path = None
if len(insights) > 0:
    fig, ax = plt.subplots(figsize=(10, max(2, len(insights)*0.4)))
    ax.axis('off')
    y_start = 1
    for i, line in enumerate(insights):
        ax.text(0, y_start - i*0.09, "- " + line, fontsize=11, ha='left', va='top', wrap=True)
    plt.title("Insights automáticos", fontsize=14, weight='bold', pad=18)
    insights_img_path = os.path.join(CARPETA_SALIDA, 'insights_img.png')
    plt.savefig(insights_img_path, bbox_inches='tight')
    plt.close()
    print(f"✅ Imagen generada: {insights_img_path}")

# ==== INICIO BLOQUE PDF COMPLETO Y EXPANDIDO ====
print("\n📄 Iniciando la generacion del informe PDF final y expandido...")
pdf_path = os.path.join(CARPETA_SALIDA, "informe_correlacion_financiera_completo.pdf")
pdf = FPDF()
pdf.set_auto_page_break(auto=True, margin=15)

# --------- Portada (Original) ---------
pdf.add_page()
pdf.set_font("helvetica", 'B', 16); pdf.cell(0, 10, "Informe de Correlación Financiera", align="C", ln=True)
pdf.set_font("helvetica", '', 12); pdf.cell(0, 10, "Autor: Leonardo Caliva", align="C", ln=True)
pdf.cell(0, 10, f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y')}", align="C", ln=True)
pdf.ln(8)
pdf.cell(0, 10, "Sitio web: https://leocaliva.com", align="C", ln=True)
pdf.ln(10)

# --------- Matriz de correlación (Original) ---------
if os.path.exists(img_path):
    pdf.add_page(); pdf.set_font("helvetica", 'B', 14); pdf.cell(0, 10, "Matriz de Correlación entre Activos", ln=True)
    pdf.image(img_path, x=15, w=180)
    pdf.ln(5); pdf.set_font("helvetica", '', 10)
    pdf.multi_cell(180, 7, encode_text("La matriz de correlación cuantifica cómo se relacionan los activos. Valores cercanos a 1 (movimiento conjunto), valores negativos (movimiento opuesto). Es clave para diversificación."))

# --------- Dendrograma (Original) y Resumen de Clusters (Nuevo) ---------
if os.path.exists(dendro_path):
    pdf.add_page(); pdf.set_font("helvetica", 'B', 14); pdf.cell(0, 10, "Dendrograma y Resumen de Clusters", ln=True)
    pdf.image(dendro_path, x=15, w=180)
    pdf.ln(5); pdf.set_font("helvetica", '', 10)
    pdf.multi_cell(180, 7, encode_text("El dendrograma agrupa activos con comportamientos similares. Las ramas cortas unen activos muy parecidos, permitiendo visualizar la estructura del portafolio."))
    pdf.ln(3)
    pdf.set_font("helvetica", 'B', 11); pdf.cell(0, 8, "Resumen de Clusters Detectados:", ln=True)
    pdf.set_font("helvetica", '', 10)
    for linea in cluster_summary_texto:
        pdf.multi_cell(0, 6, encode_text(linea))

# --------- NUEVA PÁGINA: Correlación vs Benchmark ---------
if corr_benchmark_path and os.path.exists(corr_benchmark_path):
    pdf.add_page(); pdf.set_font("helvetica", 'B', 14); pdf.cell(0, 10, f"Análisis de Sensibilidad vs Benchmark ({BENCHMARK})", ln=True)
    pdf.image(corr_benchmark_path, x=25, w=160)
    pdf.ln(5); pdf.set_font("helvetica", '', 10)
    pdf.multi_cell(180, 7, encode_text(f"Este gráfico mide la sensibilidad de cada activo al mercado (representado por {BENCHMARK}). Activos con alta correlación siguen al índice; con baja correlación, ofrecen diversificación."))
    pdf.set_font("helvetica", 'B', 11); pdf.cell(0, 8, "Insights de Sensibilidad:", ln=True)
    pdf.set_font("helvetica", '', 10)
    for linea in corr_benchmark_texto:
        pdf.multi_cell(0, 6, encode_text(linea))

# --------- NUEVA PÁGINA: Rupturas de Correlación ---------
if rupturas_recientes_texto:
    pdf.add_page(); pdf.set_font("helvetica", 'B', 14); pdf.cell(0, 10, "Análisis de Rupturas de Correlación Recientes", ln=True)
    pdf.ln(5); pdf.set_font("helvetica", '', 10)
    pdf.multi_cell(180, 7, encode_text("Este análisis identifica cambios significativos en las correlaciones entre pares de activos, lo que puede indicar cambios en las dinámicas del mercado o en la relación entre los activos."))
    pdf.set_font("helvetica", 'B', 11); pdf.cell(0, 8, "Rupturas Detectadas:", ln=True)
    pdf.set_font("helvetica", '', 10)
    for linea in rupturas_recientes_texto:
        pdf.multi_cell(0, 6, encode_text(linea))

# --------- NUEVA PÁGINA: Comparación Histórica de Correlaciones ---------
if corr_diff_path and os.path.exists(corr_diff_path):
    pdf.add_page(); pdf.set_font("helvetica", 'B', 14); pdf.cell(0, 10, "Comparación Histórica de Correlaciones (6M vs 6M Anteriores)", ln=True)
    pdf.image(corr_diff_path, x=15, w=180)
    pdf.ln(5); pdf.set_font("helvetica", '', 10)
    pdf.multi_cell(180, 7, encode_text("Este heatmap muestra las diferencias en las correlaciones entre el último período de 6 meses y el período anterior de 6 meses. Los colores indican cambios: rojo para disminución, azul para aumento."))
    pdf.set_font("helvetica", 'B', 11); pdf.cell(0, 8, "Insights de Cambios Históricos:", ln=True)
    pdf.set_font("helvetica", '', 10)
    for linea in corr_diff_texto:
        pdf.multi_cell(0, 6, encode_text(linea))

# --------- Network graph (Original) ---------
if os.path.exists(net_path):
    pdf.add_page(); pdf.set_font("helvetica", 'B', 14); pdf.cell(0, 10, "Network Graph de Correlaciones", ln=True)
    pdf.image(net_path, x=20, w=170)
    pdf.ln(5); pdf.set_font("helvetica", '', 10)
    pdf.multi_cell(180, 7, encode_text("Este gráfico muestra, como una red, los pares de activos con correlaciones fuertes (absoluta mayor a 0.3). Ayuda a detectar clústers y relaciones relevantes entre activos que pueden influir en la dinámica del portafolio."))

# --------- Distribución de correlaciones (Original) ---------
if os.path.exists(hist_path):
    pdf.add_page(); pdf.set_font("helvetica", 'B', 14); pdf.cell(0, 10, "Distribución de Correlaciones", ln=True)
    pdf.image(hist_path, x=30, w=150)
    pdf.ln(5); pdf.set_font("helvetica", '', 10)
    pdf.multi_cell(180, 7, encode_text("El histograma muestra cómo se distribuyen las correlaciones entre todos los pares de activos. Una concentración de valores cerca de cero indica buen nivel de diversificación, mientras que muchos valores altos sugieren riesgo de concentración."))

# --------- Rolling correlation (Original) ---------
for nombre, path in rolling_paths:
    if os.path.exists(path):
        pdf.add_page(); pdf.set_font("helvetica", 'B', 14); pdf.cell(0, 10, f"Rolling Correlation: {nombre}", ln=True)
        pdf.image(path, x=15, w=180)
        pdf.ln(5); pdf.set_font("helvetica", '', 10)
        pdf.multi_cell(180, 7, encode_text(f"Correlación móvil de 60 días para el par {nombre}. Sirve para ver cómo la relación entre los activos evoluciona a lo largo del tiempo."))

# --------- PCA (Original) ---------
if os.path.exists(pca_path):
    pdf.add_page(); pdf.set_font("helvetica", 'B', 14); pdf.cell(0, 10, "Varianza Explicada por PCA", ln=True)
    pdf.image(pca_path, x=30, w=150)
    pdf.ln(5); pdf.set_font("helvetica", '', 10)
    pdf.multi_cell(180, 7, encode_text("El gráfico de varianza explicada muestra cuánta información aporta cada componente principal del portafolio. Es útil para entender la dimensionalidad y la estructura interna de los retornos."))

# --------- Insights automáticos (Original) ---------
pdf.add_page()
pdf.set_font("helvetica", 'B', 15); pdf.cell(0, 12, "Insights automáticos", ln=True)
pdf.set_font("helvetica", '', 11); pdf.set_text_color(80, 80, 80)
pdf.multi_cell(0, 8, encode_text("A continuación se listan relaciones clave detectadas en la matriz de correlaciones. Estas frases resumen pares de activos que se mueven juntos (alta correlación) o en sentido opuesto (baja o negativa), lo que sirve para diversificación o cobertura.\n"))
pdf.set_text_color(0, 0, 0)
pdf.ln(2)
if insights:
    for line in insights:
        pdf.multi_cell(0, 7, encode_text("- " + line))

# --------- Clasificación por Sector (Original) ---------
if sector_img_path and os.path.exists(sector_img_path):
    pdf.add_page(); pdf.set_font("helvetica", 'B', 14); pdf.cell(0, 10, "Promedio de Correlación Intra-sectorial", ln=True)
    pdf.image(sector_img_path, x=40, w=130)
    pdf.ln(5); pdf.set_font("helvetica", '', 10)
    pdf.multi_cell(180, 7, encode_text("Esta tabla muestra la correlación promedio entre activos dentro de cada sector. Sectores con alta correlación interna pueden indicar menor diversificación dentro de ellos."))


pdf.output(pdf_path)
print(f"✅ Informe PDF completo generado: {pdf_path}")

