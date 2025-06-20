import os
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from fpdf import FPDF
from datetime import datetime
import scipy.cluster.hierarchy as sch
   

CARPETA_DATOS = './datospython1'
CARPETA_GENERAL = './datosgenerales'
CARPETA_SALIDA = './matriz_correlacion3'
os.makedirs(CARPETA_SALIDA, exist_ok=True)

print(f"Directorio actual: {os.getcwd()}")
print(f"Carpeta salida absoluta: {os.path.abspath(CARPETA_SALIDA)}")

# ==== LECTURA Y PREPARACIÓN ====
dataframes = {}
for archivo in os.listdir(CARPETA_DATOS):
    ruta = os.path.join(CARPETA_DATOS, archivo)
    nombre = os.path.splitext(archivo)[0]
    try:
        if archivo.endswith('.csv'):
            # Leer el CSV, saltando las filas 1 y 2 (índice 0 y 1) y usando la fila 0 (la que contiene 'Price', 'Close', etc.) como encabezado
            # La columna de fecha es la primera columna (índice 0) y se parsea como fecha
            df = pd.read_csv(ruta, skiprows=[1, 2], header=0, index_col=0, parse_dates=True)
        elif archivo.endswith('.xlsx') or archivo.endswith('.xls'):
            df = pd.read_excel(ruta, index_col=0, parse_dates=True)
        else:
            continue
        
        # Se busca la columna 'Close' o 'Price' (en ese orden de preferencia)
        columnas = [col for col in df.columns if col.lower() == 'close']
        if not columnas:
            columnas = [col for col in df.columns if col.lower() == 'price']
        if not columnas:
            columnas = [col for col in df.columns if col.lower() in ['adj close', 'precio_cierre']]

        if not columnas:
            continue

        df_filtrado = df[[columnas[0]]].rename(columns={columnas[0]: nombre})
        df_filtrado[nombre] = pd.to_numeric(df_filtrado[nombre], errors='coerce') # Convertir a numérico, forzando NaN en errores
        df_filtrado = df_filtrado.dropna() # Eliminar filas con NaN después de la conversión
        if not df_filtrado.empty:
            dataframes[nombre] = df_filtrado

    except Exception as e:
        print(f"⚠️ Error con {archivo}: {e}")

if not dataframes:
    print("❌ No se encontraron datos válidos.")
    exit()

df_precios = pd.concat(dataframes.values(), axis=1).dropna()
retornos = np.log(df_precios / df_precios.shift(1)).dropna()
matriz_correlacion = retornos.corr()

# ==== Excel ====
excel_path = os.path.join(CARPETA_SALIDA, 'matriz_correlacion.xlsx')
wb = Workbook()
ws = wb.active
ws.title = "Correlaciones"
for fila in dataframe_to_rows(matriz_correlacion.round(4), index=True, header=True):
    ws.append(fila)
wb.save(excel_path)

# ==== Gráfico 1: Matriz de Correlación (Heatmap) ====

# Aplicar clustering jerárquico para reordenar
linkage = sch.linkage(matriz_correlacion, method='ward')
orden = sch.dendrogram(linkage, no_plot=True)['leaves']
matriz_correlacion = matriz_correlacion.iloc[orden, orden]

# Updated Gráfico 1: Matriz de Correlación (Heatmap) con Clustering Jerárquico
img_path = os.path.join(CARPETA_SALIDA, 'heatmap_correlacion.png')
fig_height = max(10, len(matriz_correlacion) * 0.4)
fig_width = fig_height * 0.9
plt.figure(figsize=(fig_width, fig_height))
sns.set(style="whitegrid", font_scale=1.1)
# Ajustar el tamaño de la fuente de las anotaciones dinámicamente
annot_font_size = 10 if len(matriz_correlacion) < 20 else 8 if len(matriz_correlacion) < 40 else 6
# Decidir si mostrar anotaciones basado en el número de activos
annot_bool = len(matriz_correlacion) < 50 # Mostrar anotaciones solo si hay menos de 50 activos

sns.heatmap(matriz_correlacion, annot=annot_bool, fmt=".2f", cmap="coolwarm", vmin=-1, vmax=1,
            square=True, linewidths=0.5, cbar_kws={'label': 'Correlación'}, annot_kws={'size': annot_font_size})
plt.title("Matriz de Correlación entre Activos", fontsize=16, weight='bold')
plt.xticks(rotation=45, ha='right')
plt.yticks(rotation=0)
plt.tight_layout()
plt.savefig(img_path)
plt.close()
print(f"✅ Imagen generada: {img_path}")

# ==== Gráfico 2: Dendrograma ====
from scipy.cluster.hierarchy import linkage, dendrogram
dendro_path = os.path.join(CARPETA_SALIDA, 'dendrograma_clustering.png')
from scipy.spatial.distance import squareform
distance_matrix = 1 - matriz_correlacion.abs()
condensed_distance = squareform(distance_matrix, checks=False)
linkage_matrix = linkage(condensed_distance, method='ward')
plt.figure(figsize=(12, 5))
dendrogram(linkage_matrix, labels=matriz_correlacion.columns, leaf_rotation=45)
plt.title('Dendrograma de Clustering de Activos')
plt.tight_layout()
plt.savefig(dendro_path)
plt.close()
print(f"✅ Imagen generada: {dendro_path}")

# ==== Gráfico 3: Network Graph ====
import networkx as nx
net_path = os.path.join(CARPETA_SALIDA, 'network_graph.png')
plt.figure(figsize=(9, 8))
G = nx.Graph()
for activo in matriz_correlacion.columns:
    G.add_node(activo)
umbral_cor = 0.3  # Bajá este valor si querés ver más relaciones
for i, a1 in enumerate(matriz_correlacion.columns):
    for j, a2 in enumerate(matriz_correlacion.columns):
        if i < j:
            cor_val = matriz_correlacion.loc[a1, a2]
            if abs(cor_val) > umbral_cor:
                G.add_edge(a1, a2, weight=abs(cor_val))
pos = nx.spring_layout(G, seed=42)
edges = G.edges()
weights = [G[u][v]['weight']*3 for u,v in edges]
nx.draw_networkx_nodes(G, pos, node_size=800, node_color='lightblue')
nx.draw_networkx_edges(G, pos, edgelist=edges, width=weights, alpha=0.6)
nx.draw_networkx_labels(G, pos, font_size=10)
plt.title(f'Network Graph de Correlaciones (>|{umbral_cor}|)')
plt.axis('off')
plt.tight_layout()
plt.savefig(net_path)
plt.close()
print(f"✅ Imagen generada: {net_path}")

# ==== Gráfico 4: Distribución de Correlaciones ====
hist_path = os.path.join(CARPETA_SALIDA, 'correlation_hist.png')
corrs_flat = matriz_correlacion.where(~np.eye(matriz_correlacion.shape[0], dtype=bool)).stack()
plt.figure(figsize=(8, 5))
sns.histplot(corrs_flat, kde=True, bins=20, color='slateblue')
plt.title('Distribución y densidad de correlaciones (Pearson)')
plt.xlabel('Correlación')
plt.ylabel('Frecuencia')
plt.tight_layout()
plt.savefig(hist_path)
plt.close()
print(f"✅ Imagen generada: {hist_path}")

# ==== Gráfico 5: Rolling Correlation (top 2 pares) ====
rolling_paths = []
corrs = matriz_correlacion.where(~np.eye(matriz_correlacion.shape[0], dtype=bool)).abs().unstack().sort_values(ascending=False)
top_pairs = []
for (a1, a2) in corrs.index:
    if a1 != a2 and (a2, a1) not in top_pairs:
        top_pairs.append((a1, a2))
    if len(top_pairs) == 2:
        break
ROLLING_WINDOW = 60
for a1, a2 in top_pairs:
    roll_path = os.path.join(CARPETA_SALIDA, f'rollingcorr_{a1}_{a2}.png')
    serie = retornos[a1].rolling(ROLLING_WINDOW).corr(retornos[a2])
    plt.figure(figsize=(10, 4))
    plt.plot(serie.index, serie, label=f'{a1}-{a2}')
    plt.axhline(0, color='gray', ls='--')
    plt.title(f'Correlación Rolling 60 días: {a1}-{a2}')
    plt.ylabel('Correlación')
    plt.xlabel('Fecha')
    plt.legend()
    plt.tight_layout()
    plt.savefig(roll_path)
    plt.close()
    rolling_paths.append((f"{a1}-{a2}", roll_path))
    print(f"✅ Imagen generada: {roll_path}")

# ==== Gráfico 6: PCA Varianza Explicada ====
from sklearn.decomposition import PCA
pca_path = os.path.join(CARPETA_SALIDA, 'pca_varianza.png')
retornos_for_pca = retornos.dropna(axis=1, how='any')
pca_ok = False
if retornos_for_pca.shape[0] > 0:
    pca = PCA()
    pca.fit(retornos_for_pca)
    explained_var = pca.explained_variance_ratio_
    plt.figure(figsize=(8, 4))
    plt.bar(range(1, len(explained_var)+1), explained_var*100)
    plt.xlabel('Componente Principal')
    plt.ylabel('% de Varianza Explicada')
    plt.title('Varianza Explicada por PCA')
    plt.tight_layout()
    plt.savefig(pca_path)
    plt.close()
    pca_ok = True
    print(f"✅ Imagen generada: {pca_path}")

# ==== INSIGHTS AUTOMÁTICOS ====
insights = []
for i in range(len(matriz_correlacion.columns)):
    for j in range(i+1, len(matriz_correlacion.columns)):
        activo1 = matriz_correlacion.columns[i]
        activo2 = matriz_correlacion.columns[j]
        correlacion = matriz_correlacion.iloc[i, j]
        if correlacion >= 0.8:
            insights.append(f"Alta correlación positiva entre {activo1} y {activo2} ({correlacion:.2f}). Se mueven casi en sincronía.")
        elif correlacion <= -0.5:
            insights.append(f"Correlación negativa entre {activo1} y {activo2} ({correlacion:.2f}). Tienden a moverse en direcciones opuestas.")
        elif -0.2 <= correlacion <= 0.2:
            insights.append(f"Baja correlación entre {activo1} y {activo2} ({correlacion:.2f}). Buena combinación para diversificación.")

insight_file = os.path.join(CARPETA_SALIDA, 'insights_correlacion.txt')
with open(insight_file, 'w', encoding='utf-8') as f:
    f.write("Insights Cuantitativos sobre Correlaciones\n")
    f.write("="*50 + "\n\n")
    for linea in insights:
        f.write(linea + "\n")

# ==== CLASIFICACIÓN POR SECTOR ====
sectores_path = os.path.join(CARPETA_GENERAL, 'sectores.csv')
tabla_sector = ""
sector_correlations = []
mapa_sectores = {}
if os.path.exists(sectores_path):
    df_sectores = pd.read_csv(sectores_path)
    df_sectores.columns = df_sectores.columns.str.strip()
    posibles_col_activo = ['Activo', 'Ticker', 'Symbol']
    posibles_col_sector = ['Sector', 'GICS Sector']
    col_activo = next((col for col in posibles_col_activo if col in df_sectores.columns), None)
    col_sector = next((col for col in posibles_col_sector if col in df_sectores.columns), None)
    if col_activo and col_sector:
        mapa_sectores = dict(zip(df_sectores[col_activo], df_sectores[col_sector]))
        activos_por_sector = {}
        for activo in matriz_correlacion.columns:
            sector = mapa_sectores.get(activo, 'Sin clasificar')
            activos_por_sector.setdefault(sector, []).append(activo)
        for sector, activos in activos_por_sector.items():
            if len(activos) < 2:
                continue
            submatriz = matriz_correlacion.loc[activos, activos]
            cor_media = submatriz.where(~np.eye(len(submatriz), dtype=bool)).mean().mean()
            sector_correlations.append((sector, round(cor_media, 3)))
        tabla_sector += "\nPromedio de correlación intra-sectorial:\n"
        for sector, media in sector_correlations:
            tabla_sector += f" - {sector}: {media}\n"
    else:
        tabla_sector = "\n⚠️ El archivo de sectores no contiene las columnas necesarias para clasificar.\n"
else:
    tabla_sector = "\n📂 No se encontró el archivo 'sectores.csv'.\n"

# ==== EXPORTAR SECTORIAL A EXCEL ====
with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
    if sector_correlations:
        df_sector = pd.DataFrame(sector_correlations, columns=['Sector', 'Correlacion Media'])
        df_sector.to_excel(writer, sheet_name='Correlacion_Sector', index=False)
    if mapa_sectores:
        df_map = pd.DataFrame(list(mapa_sectores.items()), columns=['Activo', 'Sector'])
        df_map.to_excel(writer, sheet_name='Activos_Sector', index=False)
    df_insights = pd.DataFrame(insights, columns=['Insight'])
    df_insights.to_excel(writer, sheet_name='Insights', index=False)

# ==== IMAGEN DE TABLA SECTORIAL ====
sector_img_path = None
if sector_correlations:
    df_sector = pd.DataFrame(sector_correlations, columns=['Sector', 'Correlación Media'])
    fig, ax = plt.subplots(figsize=(6, max(2, len(df_sector)*0.5)))
    ax.axis('tight')
    ax.axis('off')
    table = ax.table(cellText=df_sector.values,
                     colLabels=df_sector.columns,
                     loc='center',
                     cellLoc='center')
    plt.title("Promedio de correlación intra-sectorial", fontsize=14, weight='bold')
    sector_img_path = os.path.join(CARPETA_SALIDA, 'sector_correlation.png')
    plt.savefig(sector_img_path, bbox_inches='tight')
    plt.close()
    print(f"✅ Imagen generada: {sector_img_path}")

# ==== IMAGEN DE INSIGHTS AUTOMÁTICOS ====
insights_img_paths = [] # Initialize here
if len(insights) > 0:
    chunk_size = 30  # máximo de líneas por imagen
    for chunk_index, start in enumerate(range(0, len(insights), chunk_size)):
        chunk = insights[start:start + chunk_size]
        fig, ax = plt.subplots(figsize=(10, max(2, len(chunk)*0.4)))
        ax.axis('off')
        y_start = 1
        for i, line in enumerate(chunk):
            ax.text(0, y_start - i*0.09, "- " + line, fontsize=11, ha='left', va='top', wrap=True)
        plt.title(f"Insights automáticos (parte {chunk_index + 1})", fontsize=14, weight='bold', pad=18)
        chunk_path = os.path.join(CARPETA_SALIDA, f'insights_img_{chunk_index + 1}.png')
        plt.savefig(chunk_path, bbox_inches='tight')
        plt.close()
        insights_img_paths.append(chunk_path)
        plt.close()
        print(f"✅ Imagen generada: {chunk_path}")

# ==== INICIO BLOQUE PDF ====
# ==== PDF PROFESIONAL COMPLETO ====
pdf_path = os.path.join(CARPETA_SALIDA, "informe_correlacion_financiera.pdf")
pdf = FPDF()
pdf.set_auto_page_break(auto=True, margin=15)
pdf.add_page()

# --------- Portada ---------
pdf.set_font("times", "B", 16)
pdf.cell(0, 10, "Informe de Correlación Financiera", align="C")
pdf.ln(10)
pdf.set_font("helvetica", '', 12)
pdf.cell(0, 10, "Autor: Leonardo Caliva", align="C")
pdf.ln(8)
pdf.cell(0, 10, f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y')}", align="C")
pdf.ln(8)
pdf.cell(0, 10, "Sitio web: https://leocaliva.com", align="C")
pdf.ln(10)

# --------- Matriz de correlación ---------
if os.path.exists(img_path):
    pdf.set_font("helvetica", 'B', 14)
    pdf.cell(0, 10, "Matriz de Correlación entre Activos")
    pdf.ln(8)
    pdf.image(img_path, x=15, w=180)
    pdf.ln(5)
    pdf.set_font("helvetica", '', 10)
    pdf.multi_cell(180, 7,
        "¿Qué muestra este gráfico?\n"
        "La matriz de correlación cuantifica cómo se relacionan los activos entre sí. "
        "Valores cercanos a 1 indican que los activos suelen moverse juntos, mientras que valores negativos muestran que tienden a moverse en direcciones opuestas. "
        "Esta herramienta es fundamental para detectar riesgos de concentración y oportunidades de diversificación."
    )

# --------- Dendrograma ---------
if os.path.exists(dendro_path):
    pdf.add_page()
    pdf.set_font("helvetica", 'B', 14)
    pdf.cell(0, 10, "Dendrograma de Clustering")
    pdf.ln(8)
    pdf.image(dendro_path, x=15, w=180)
    pdf.ln(5)
    pdf.set_font("helvetica", '', 10)
    pdf.multi_cell(180, 7,
        "¿Para qué sirve?\n"
        "El dendrograma agrupa los activos que se comportan de forma similar en clusters jerárquicos. "
        "Permite visualizar patrones de comportamiento y segmentar el portafolio según similitudes."
    )

# --------- Network graph ---------
if os.path.exists(net_path):
    pdf.add_page()
    pdf.set_font("helvetica", 'B', 14)
    pdf.cell(0, 10, "Network Graph de Correlaciones")
    pdf.ln(8)
    pdf.image(net_path, x=20, w=170)
    pdf.ln(5)
    pdf.set_font("helvetica", '', 10)
    pdf.multi_cell(180, 7,
        "¿Para qué sirve?\n"
        "Este gráfico muestra, como una red, los pares de activos con correlaciones fuertes (absoluta mayor a 0.3). "
        "Ayuda a detectar clústers y relaciones relevantes entre activos que pueden influir en la dinámica del portafolio."
    )

# --------- Distribución de correlaciones ---------
if os.path.exists(hist_path):
    pdf.add_page()
    pdf.set_font("helvetica", 'B', 14)
    pdf.cell(0, 10, "Distribución de Correlaciones")
    pdf.ln(8)
    pdf.image(hist_path, x=30, w=150)
    pdf.ln(5)
    pdf.set_font("helvetica", '', 10)
    pdf.multi_cell(180, 7,
        "¿Qué significa?\n"
        "El histograma muestra cómo se distribuyen las correlaciones entre todos los pares de activos. "
        "Una concentración de valores cerca de cero indica buen nivel de diversificación, mientras que muchos valores altos sugieren riesgo de concentración."
    )

# --------- Rolling correlation ---------
for nombre, path in rolling_paths:
    if os.path.exists(path):
        pdf.add_page()
        pdf.set_font("helvetica", 'B', 14)
        pdf.cell(0, 10, f"Rolling Correlation: {nombre}")
        pdf.ln(8)
        pdf.image(path, x=15, w=180)
        pdf.ln(5)
        pdf.set_font("helvetica", '', 10)
        pdf.multi_cell(180, 7,
            f"¿Qué muestra?\n"
            f"Correlación móvil de 60 días para el par {nombre}. Sirve para ver cómo la relación entre los activos evoluciona a lo largo del tiempo."
        )

# --------- PCA ---------
if pca_ok and os.path.exists(pca_path):
    pdf.add_page()
    pdf.set_font("helvetica", 'B', 14)
    pdf.cell(0, 10, "Varianza Explicada por PCA")
    pdf.ln(8)
    pdf.image(pca_path, x=30, w=150)
    pdf.ln(5)
    pdf.set_font("helvetica", '', 10)
    pdf.multi_cell(180, 7,
        "¿Qué significa?\n"
        "El gráfico de varianza explicada muestra cuánta información aporta cada componente principal del portafolio. "
        "Es útil para entender la dimensionalidad y la estructura interna de los retornos."
    )

# --------- Insights automáticos (texto mejorado) ---------
pdf.add_page()
pdf.set_font("helvetica", 'B', 14)
pdf.cell(0, 10, "Insights Cuantitativos sobre Correlaciones")
pdf.ln(8)
for path in insights_img_paths:
    if os.path.exists(path):
        pdf.image(path, x=15, w=180)
        pdf.ln(5)

# --------- Clasificación por sector ---------
if sector_img_path and os.path.exists(sector_img_path):
    pdf.add_page()
    pdf.set_font("helvetica", 'B', 14)
    pdf.cell(0, 10, "Correlación por Sector")
    pdf.ln(8)
    pdf.image(sector_img_path, x=15, w=180)
    pdf.ln(5)
    pdf.set_font("helvetica", '', 10)
    pdf.multi_cell(180, 7,
        "¿Qué muestra?\n"
        "Este gráfico resume la correlación promedio dentro de cada sector, "
        "lo que es útil para entender la cohesión de los activos dentro de una misma industria."
    )

# Final del PDF
try:
    pdf.output(pdf_path)
    print(f"✅ PDF profesional generado en: {pdf_path}")
except Exception as e:
    print(f"❌ Error al generar el PDF: {e}")




