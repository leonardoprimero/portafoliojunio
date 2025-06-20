#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Análisis Financiero Final - TODOS LOS PROBLEMAS CORREGIDOS
=========================================================
Análisis financiero profesional desarrollado por Leonardo Caliva.
Versión final con gráficos que NO se superponen con pie de página
y Excel con imágenes REALMENTE separadas de las tablas.

Autor: Leonardo Caliva
Portfolio: leocaliva.com
"""

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from matplotlib.backends.backend_pdf import PdfPages
import shutil
from datetime import datetime, timedelta
from textwrap import wrap
import warnings
warnings.filterwarnings("ignore")

# ========================= CONFIGURACIÓN =========================

# CONFIGURACIÓN PRINCIPAL

# Escala de los gráficos de precios
# Opciones: "linear" o "log"
ESCALA_PRECIOS = "linear"

# Tema visual de los gráficos
# Opciones:
# "light" (fondo claro, paleta profesional)
# "dark" (fondo oscuro, paleta profesional)
# "vintage" (estilo avejentado, paleta profesional)
# "normal" (estilo por defecto de seaborn-v0_8-whitegrid)
TEMA_GRAFICOS = "dark"

# Medias Móviles (días)
SMA_CORTA = 21
SMA_MEDIA = 63
SMA_LARGA = 252

# Configuración de indicadores técnicos adicionales
ACTIVAR_RSI = True # Cambiar a True para activar el cálculo y visualización
RSI_PERIOD = 14

ACTIVAR_MACD = True # Cambiar a True para activar el cálculo y visualización
MACD_FAST_PERIOD = 12
MACD_SLOW_PERIOD = 26
MACD_SIGNAL_PERIOD = 9

ACTIVAR_BBANDS = True # Cambiar a True para activar el cálculo y visualización
BBANDS_PERIOD = 20
BBANDS_DEV = 2

ACTIVAR_VOLUMEN = True # Cambiar a True para activar el análisis de volumen

# Configuración de archivos y carpetas
CARPETA_DATOS = './datospython1'  # Carpeta donde están los datos reales (CSV)
CARPETA_GRAFICOS = './analisisPorActivo'  # Carpeta donde se guardarán los gráficos generados
ARCHIVO_PDF = os.path.join(CARPETA_GRAFICOS, 'reporte_analisis_final.pdf')
ARCHIVO_EXCEL = os.path.join(CARPETA_GRAFICOS, 'analisis_activos_final.xlsx')


# Información del autor
AUTOR_NOMBRE = "Leonardo Caliva"
AUTOR_PORTFOLIO = "leocaliva.com"

# Configuración de matplotlib para A4 y sin superposición
def configurar_matplotlib(tema):
    """Configurar matplotlib para A4 perfecto y aplicar tema"""
    # Colores base para los gráficos según el tema
    if tema == "light":
        plt.style.use('seaborn-v0_8-whitegrid')
        sns.set_palette("viridis") # Paleta moderna y profesional
        facecolor_plot = 'white'
        textcolor_plot = '#212121'
        linecolor_plot = '#424242'
    elif tema == "dark":
        plt.style.use('dark_background') # Fondo oscuro
        sns.set_palette("viridis") # Paleta moderna y profesional
        facecolor_plot = '#282c34'
        textcolor_plot = '#abb2bf'
        linecolor_plot = '#61afef'
    elif tema == "vintage":
        plt.style.use('seaborn-v0_8-pastel') # Estilo más suave, vintage
        sns.set_palette("deep") # Paleta con colores más apagados
        facecolor_plot = '#f0f0d0'
        textcolor_plot = '#5c5c5c'
        linecolor_plot = '#8b4513'
    else: # "normal" o cualquier otro valor
        plt.style.use('seaborn-v0_8-whitegrid')
        sns.set_palette("Set2")
        facecolor_plot = 'white'
        textcolor_plot = '#212121'
        linecolor_plot = '#424242'

    plt.rcParams.update({
        'figure.figsize': (7.5, 5),  # MÁS PEQUEÑO para A4
        'axes.labelsize': 9,
        'axes.titlesize': 11,
        'legend.fontsize': 8,
        'xtick.labelsize': 8,
        'ytick.labelsize': 8,
        'font.size': 9,
        'pdf.fonttype': 42,
        'ps.fonttype': 42,
        'figure.facecolor': 'white',  # Fondo de la figura siempre blanco para el PDF
        'axes.facecolor': facecolor_plot, # Fondo del área de plot según el tema
        'text.color': textcolor_plot,
        'axes.labelcolor': textcolor_plot,
        'xtick.color': textcolor_plot, 
        'ytick.color': textcolor_plot, 
        'grid.color': linecolor_plot,
        'grid.alpha': 0.3
    })
    
    # Configurar fuentes para caracteres griegos
    # Asegúrate de que las fuentes estén en la carpeta 'fonts' en el mismo directorio que el script
    try:
        from matplotlib import font_manager
        font_dirs = [os.path.join(os.path.dirname(__file__), 'fonts')]
        font_files = font_manager.findSystemFonts(fontpaths=font_dirs)
        for font_file in font_files:
            font_manager.fontManager.addfont(font_file)
        
        # Establecer la fuente predeterminada para el texto
        plt.rcParams['font.family'] = 'DejaVu Sans'
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
        
        # Para negritas
        plt.rcParams['font.weight'] = 'normal'
        plt.rcParams['axes.titleweight'] = 'bold'
        plt.rcParams['axes.labelweight'] = 'normal'
        
    except Exception as e:
        print(f"Advertencia: No se pudieron cargar las fuentes personalizadas. {e}")
        print("Asegúrate de que los archivos de fuente (.ttf) estén en la carpeta 'fonts'.")

# ========================= ANÁLISIS FINANCIERO =========================

def crear_portada_pdf(pdf, tickers, autor_nombre, autor_portfolio):
    """Crea una portada profesional para el PDF"""
    fig = plt.figure(figsize=(8.27, 11.69))  # Tamaño A4
    
    # Configurar el layout sin ejes
    ax = fig.add_subplot(111)
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis('off')
    
    # Encabezado principal
    ax.text(0.5, 0.85, 'ANÁLISIS FINANCIERO', 
            fontsize=24, weight='bold', ha='center', va='center',
            color='#1a237e', transform=ax.transAxes)
    
    # Línea decorativa
    ax.plot([0.1, 0.9], [0.82, 0.82], color='#1a237e', linewidth=2, transform=ax.transAxes)
    
    # Armar texto de tickers con ajuste automático
    tickers_str = ", ".join(tickers)

    # Elegir fuente según largo
    if len(tickers_str) <= 80:
        font_size = 14
    elif len(tickers_str) <= 150:
        font_size = 12
    elif len(tickers_str) <= 250:
        font_size = 10
    else:
        font_size = 9

    # Envolver para que no se pase del ancho y ajustar posición Y
    wrapped_lines = wrap(tickers_str, width=80)
    # Calcular la posición Y inicial para centrar el bloque de tickers
    num_lines = len(wrapped_lines)
    y_pos_start = 0.78 + (num_lines * 0.0175) # Ajuste para centrar visualmente
    
    for i, line in enumerate(wrapped_lines):
        ax.text(0.5, y_pos_start - (i * 0.035), line, # Ajuste vertical más fino
                fontsize=font_size, ha='center', va='center', style='italic',
                color='#424242', transform=ax.transAxes)
    
    # Posición inicial de la descripción ajustada para que no se superponga
    descripcion_y_start = y_pos_start - (num_lines * 0.035) - 0.05 # Deja espacio después de los tickers
    
    descripcion = """CONTENIDO DEL INFORME:

• Análisis de precios históricos con medias móviles (21d, 63d, 252d)
• Cálculo de retornos diarios y métricas de volatilidad
• Distribución estadística de rendimientos
• Análisis técnico automatizado por activo
• Comparación de rendimientos entre activos
• Recomendaciones técnicas basadas en indicadores

METODOLOGÍA:

• Medias móviles simples para identificación de tendencias
• Análisis de volatilidad para evaluación de riesgo
• Cálculo de retornos anualizados con base en 252 días hábiles
• Análisis estadístico de distribución de retornos"""
    
    ax.text(0.1, descripcion_y_start, descripcion, # Usa la posición Y ajustada
            fontsize=11, ha='left', va='top',
            color='#424242', transform=ax.transAxes)
    
    # Información del autor
    ax.text(0.5, 0.25, f'Desarrollado por: {autor_nombre}', 
            fontsize=14, weight='bold', ha='center', va='center',
            color='#1a237e', transform=ax.transAxes)
    
    ax.text(0.5, 0.21, f'{autor_portfolio}', 
            fontsize=12, ha='center', va='center', style='italic',
            color='#424242', transform=ax.transAxes)
    
    # Fecha de generación
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    ax.text(0.5, 0.15, f'Fecha de análisis: {fecha_actual}', 
            fontsize=11, ha='center', va='center',
            color='#424242', transform=ax.transAxes)
    
    # Espacio reservado - sin nota adicional para mantener diseño limpio
    
    # Pie de página (sin interferir con contenido)
    fig.text(0.5, 0.02, autor_portfolio, 
            ha='center', va='bottom', fontsize=9, color='#1a237e')
    
    plt.tight_layout()
    pdf.savefig(fig, bbox_inches='tight')
    plt.close(fig)

def agregar_pie_pagina_correcto(fig, numero_pagina, total_paginas, autor_nombre, autor_portfolio):
    """Agrega pie de página profesional sin superponerse"""
    # Ajusta estas coordenadas y tamaño de fuente según sea necesario para tu diseño A4.
    # El 0.5 es el centro horizontal. El 0.02 es la distancia desde la parte inferior.
    fig.text(0.5, 0.02, 
             f'Página {numero_pagina}/{total_paginas} | {autor_nombre} | {autor_portfolio} | {datetime.now().year}',
             ha='center', va='bottom', fontsize=8, color='#666666')


def crear_grafico_precios(df, nombre, escala, numero_pagina=1, total_paginas=1, autor_nombre=AUTOR_NOMBRE, autor_portfolio=AUTOR_PORTFOLIO, sma_corta=SMA_CORTA, sma_media=SMA_MEDIA, sma_larga=SMA_LARGA):
    """Crea gráfico de precios SIN superposición con pie de página"""
    # TAMAÑO REDUCIDO para que quepa perfecto en A4
    fig, ax = plt.subplots(figsize=(7.5, 4.5))  
    
    # Configurar escala
    if escala == "log":
        ax.set_yscale('log')
        titulo_extra = " (Escala Logarítmica)"
    else:
        titulo_extra = " (Escala Linear)"
    
    # Plotear datos
    df['Precio'].plot(ax=ax, label='Precio', linewidth=1.5, color='#1976d2')
    df[f'SMA_{sma_corta}'].plot(ax=ax, label=f'SMA {sma_corta}d', linewidth=1, color='#d32f2f')
    df[f'SMA_{sma_media}'].plot(ax=ax, label=f'SMA {sma_media}d', linewidth=1, color='#f57c00')
    df[f'SMA_{sma_larga}'].plot(ax=ax, label=f'SMA {sma_larga}d', linewidth=1.5, color='#388e3c')
    
    # Plotear Bandas de Bollinger si están activadas
    if ACTIVAR_BBANDS and 'BB_Upper' in df.columns and 'BB_Lower' in df.columns and 'BB_Middle' in df.columns:
        ax.plot(df['BB_Upper'], label='Banda Superior BB', color='gray', linestyle='--', linewidth=0.8)
        ax.plot(df['BB_Middle'], label='Banda Media BB', color='gray', linestyle='-', linewidth=0.8)
        ax.plot(df['BB_Lower'], label='Banda Inferior BB', color='gray', linestyle='--', linewidth=0.8)

    ax.set_title(f'{nombre} - Análisis de Precios{titulo_extra}', 
                fontsize=12, weight='bold', color='#1a237e')
    ax.set_ylabel("Precio (USD)", fontsize=10)
    ax.set_xlabel("Fecha", fontsize=10)
    ax.legend(loc='upper left', fontsize=8)
    ax.grid(True, alpha=0.3)
    
    # CRUCIAL: Dejar espacio para pie de página
    plt.tight_layout()
    plt.subplots_adjust(bottom=0.15)  # Espacio para pie de página
    
    # Agregar pie de página SIN superposición
    agregar_pie_pagina_correcto(fig, numero_pagina, total_paginas, autor_nombre, autor_portfolio)
    
    return fig

def crear_grafico_rsi(df, nombre, numero_pagina=1, total_paginas=1, autor_nombre=AUTOR_NOMBRE, autor_portfolio=AUTOR_PORTFOLIO):
    """Crea gráfico de RSI"""
    fig, ax = plt.subplots(figsize=(7.5, 2.5)) # Más pequeño para RSI
    if 'RSI' in df.columns:
        ax.plot(df['RSI'], label='RSI', color='#8e24aa', linewidth=1.2)
        ax.axhline(70, linestyle='--', alpha=0.6, color='red', label='Sobrecompra (70)')
        ax.axhline(30, linestyle='--', alpha=0.6, color='green', label='Sobreventa (30)')
        ax.set_ylim(0, 100)
        ax.set_title(f'{nombre} - Índice de Fuerza Relativa (RSI)', fontsize=12, weight='bold', color='#1a237e')
        ax.set_ylabel('RSI', fontsize=10)
        ax.set_xlabel('Fecha', fontsize=10)
        ax.legend(loc='upper left', fontsize=8)
        ax.grid(True, alpha=0.3)
    else:
        ax.text(0.5, 0.5, 'RSI no disponible', horizontalalignment='center', verticalalignment='center', transform=ax.transAxes, fontsize=12, color='gray')
        ax.axis('off')

    plt.tight_layout()
    plt.subplots_adjust(bottom=0.15)
    agregar_pie_pagina_correcto(fig, numero_pagina, total_paginas, autor_nombre, autor_portfolio)
    return fig

def crear_grafico_macd(df, nombre, numero_pagina=1, total_paginas=1, autor_nombre=AUTOR_NOMBRE, autor_portfolio=AUTOR_PORTFOLIO):
    """Crea gráfico de MACD"""
    fig, ax = plt.subplots(figsize=(7.5, 2.5)) # Más pequeño para MACD
    if 'MACD' in df.columns and 'MACD_Signal' in df.columns and 'MACD_Hist' in df.columns:
        ax.plot(df['MACD'], label='MACD', color='#00796b', linewidth=1.2)
        ax.plot(df['MACD_Signal'], label='Señal', color='#d84315', linewidth=1.2)
        ax.bar(df.index, df['MACD_Hist'], label='Histograma', color='#4db6ac', alpha=0.6)
        ax.axhline(0, color='gray', linestyle='--', linewidth=0.8)
        ax.set_title(f'{nombre} - MACD', fontsize=12, weight='bold', color='#1a237e')
        ax.set_ylabel('Valor', fontsize=10)
        ax.set_xlabel('Fecha', fontsize=10)
        ax.legend(loc='upper left', fontsize=8)
        ax.grid(True, alpha=0.3)
    else:
        ax.text(0.5, 0.5, 'MACD no disponible', horizontalalignment='center', verticalalignment='center', transform=ax.transAxes, fontsize=12, color='gray')
        ax.axis('off')

    plt.tight_layout()
    plt.subplots_adjust(bottom=0.15)
    agregar_pie_pagina_correcto(fig, numero_pagina, total_paginas, autor_nombre, autor_portfolio)
    return fig

def crear_grafico_volumen(df, nombre, numero_pagina=1, total_paginas=1, autor_nombre=AUTOR_NOMBRE, autor_portfolio=AUTOR_PORTFOLIO):
    """Crea gráfico de Volumen"""
    fig, ax = plt.subplots(figsize=(7.5, 2.5)) # Más pequeño para Volumen
    if 'Volume' in df.columns:
        ax.bar(df.index, df['Volume'], color='#6a1b9a', alpha=0.7)
        ax.set_title(f'{nombre} - Volumen de Negociación', fontsize=12, weight='bold', color='#1a237e')
        ax.set_ylabel('Volumen', fontsize=10)
        ax.set_xlabel('Fecha', fontsize=10)
        ax.grid(True, alpha=0.3)
    else:
        ax.text(0.5, 0.5, 'Volumen no disponible', horizontalalignment='center', verticalalignment='center', transform=ax.transAxes, fontsize=12, color='gray')
        ax.axis('off')

    plt.tight_layout()
    plt.subplots_adjust(bottom=0.15)
    agregar_pie_pagina_correcto(fig, numero_pagina, total_paginas, autor_nombre, autor_portfolio)
    return fig

def crear_grafico_retornos(df, nombre, numero_pagina=1, total_paginas=1, autor_nombre=AUTOR_NOMBRE, autor_portfolio=AUTOR_PORTFOLIO):
    """Crea gráfico de retornos SIN superposición"""
    fig, ax = plt.subplots(figsize=(7.5, 4))
    
    df['Retorno'].plot(ax=ax, color='#ff6f00', alpha=0.7, linewidth=0.8)
    ax.axhline(0, color='#424242', linestyle='--', linewidth=1, alpha=0.8)
    
    # Agregar bandas de volatilidad
    retorno_medio = df['Retorno'].mean()
    vol_std = df['Retorno'].std()
    ax.axhline(retorno_medio + vol_std, color='#d32f2f', linestyle=':', alpha=0.6, label=f'+1σ: {retorno_medio + vol_std:.4f}')
    ax.axhline(retorno_medio - vol_std, color='#d32f2f', linestyle=':', alpha=0.6, label=f'-1σ: {retorno_medio - vol_std:.4f}')
    
    ax.set_title(f'{nombre} - Retornos Diarios', fontsize=12, weight='bold', color='#1a237e')
    ax.set_ylabel("Retorno (%)", fontsize=10)
    ax.set_xlabel("Fecha", fontsize=10)
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)
    
    # Espacio para pie de página
    plt.tight_layout()
    plt.subplots_adjust(bottom=0.15)
    
    agregar_pie_pagina_correcto(fig, numero_pagina, total_paginas, autor_nombre, autor_portfolio)
    
    return fig

def crear_histograma_retornos(df, nombre, numero_pagina=1, total_paginas=1, autor_nombre=AUTOR_NOMBRE, autor_portfolio=AUTOR_PORTFOLIO):
    """Crea histograma SIN superposición"""
    fig, ax = plt.subplots(figsize=(7.5, 4.5))
    
    retornos_limpios = df['Retorno'].dropna()
    
    # Histograma con curva de densidad
    sns.histplot(retornos_limpios, bins=50, kde=True, ax=ax, 
                color='#00695c', alpha=0.7, stat='density')
    
    # Estadísticas
    mu = retornos_limpios.mean()
    sigma = retornos_limpios.std()
    
    # Líneas de referencia
    ax.axvline(mu, color='#d32f2f', linestyle='--', linewidth=2, label=f'Media: {mu:.4f}')
    ax.axvline(mu + sigma, color='#388e3c', linestyle='--', linewidth=1.5, label=f'+1σ: {mu + sigma:.4f}')
    ax.axvline(mu - sigma, color='#388e3c', linestyle='--', linewidth=1.5, label=f'-1σ: {mu - sigma:.4f}')
    
    ax.set_title(f'{nombre} - Distribución de Retornos', fontsize=12, weight='bold', color='#1a237e')
    ax.set_xlabel("Retorno Diario", fontsize=10)
    ax.set_ylabel("Densidad", fontsize=10)
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)
    
    # Espacio para pie de página
    plt.tight_layout()
    plt.subplots_adjust(bottom=0.15)
    
    agregar_pie_pagina_correcto(fig, numero_pagina, total_paginas, autor_nombre, autor_portfolio)
    
    return fig

def crear_comentario_activo(nombre, stats, df, numero_pagina=1, total_paginas=1, autor_nombre=AUTOR_NOMBRE, autor_portfolio=AUTOR_PORTFOLIO, sma_corta=SMA_CORTA, sma_media=SMA_MEDIA, sma_larga=SMA_LARGA):
    """Crea página de comentarios para cada activo"""
    fig, ax = plt.subplots(figsize=(8.27, 10.5))  # Reducido para dejar espacio al pie
    ax.axis('off')
    
    # Título del activo
    ax.text(0.5, 0.95, f'ANÁLISIS TÉCNICO: {nombre}', 
            fontsize=16, weight='bold', ha='center', va='top',
            color='#1a237e', transform=ax.transAxes)
    
    # Línea decorativa
    ax.plot([0.1, 0.9], [0.92, 0.92], color='#1a237e', linewidth=1.5, transform=ax.transAxes)
    
    # Métricas principales
    metricas_texto = f"""MÉTRICAS FINANCIERAS PRINCIPALES:

Precio Actual: ${stats['precio_actual']:.2f} USD
Retorno Anual Estimado: {stats['retorno_anual']*100:.2f}%
Volatilidad Diaria: {stats['volatilidad_diaria']*100:.2f}%
Volatilidad Mensual: {stats['volatilidad_mensual']*100:.2f}%

MEDIAS MÓVILES:
• SMA {sma_corta} días: ${stats[f'sma_{sma_corta}']:.2f}
• SMA {sma_media} días: ${stats[f'sma_{sma_media}']:.2f} 
• SMA {sma_larga} días: ${stats[f'sma_{sma_larga}']:.2f}

ANÁLISIS TÉCNICO:"""
    
    # Análisis de tendencia
    precio_actual = stats['precio_actual']
    sma_larga_val = stats[f'sma_{sma_larga}']
    sma_media_val = stats[f'sma_{sma_media}']
    sma_corta_val = stats[f'sma_{sma_corta}']
    
    if precio_actual > sma_larga_val:
        tendencia = "ALCISTA - El precio está por encima de la media móvil anual"
        tendencia_simbolo = "↗"
    else:
        tendencia = "BAJISTA - El precio está por debajo de la media móvil anual"
        tendencia_simbolo = "↘"
    
    if sma_corta_val > sma_media_val > sma_larga_val:
        momentum = "MOMENTUM POSITIVO - Todas las medias en orden ascendente"
        momentum_simbolo = "↑"
    elif sma_corta_val < sma_media_val < sma_larga_val:
        momentum = "MOMENTUM NEGATIVO - Todas las medias en orden descendente"
        momentum_simbolo = "↓"
    else:
        momentum = "MOMENTUM MIXTO - Medias móviles entrelazadas"
        momentum_simbolo = "→"
    
    # Análisis de volatilidad
    vol_diaria = stats['volatilidad_diaria']
    if vol_diaria < 0.015:
        vol_nivel = "BAJA"
    elif vol_diaria < 0.025:
        vol_nivel = "MODERADA"
    else:
        vol_nivel = "ALTA"
    
    analisis_texto = f"""
{tendencia_simbolo} Tendencia: {tendencia}

{momentum_simbolo} Momentum: {momentum}

• Volatilidad: {vol_nivel} ({vol_diaria*100:.2f}% diaria)

INTERPRETACIÓN TÉCNICA:

La evaluación técnica de {nombre} se basa en el análisis de medias móviles 
y métricas de volatilidad. El precio actual de ${precio_actual:.2f} en relación 
con las medias móviles indica la dirección predominante del activo.

RECOMENDACIONES TÉCNICAS:

• Soporte técnico estimado: ${min(sma_media_val, sma_larga_val):.2f}
• Resistencia técnica estimada: ${max(sma_corta_val, precio_actual * 1.05):.2f}
• Nivel de riesgo: {vol_nivel}

CONSIDERACIONES:

El análisis se basa en datos históricos y metodologías de análisis técnico 
reconocidas. Las condiciones del mercado pueden cambiar rápidamente y 
requieren monitoreo constante."""
    
    # Añadir información de indicadores técnicos si están activados
    indicadores_texto = ""
    if ACTIVAR_RSI and 'rsi_actual' in stats and pd.notna(stats['rsi_actual']):
        indicadores_texto += f"\n• RSI ({RSI_PERIOD}d): {stats['rsi_actual']:.2f} (Sobrec.: >70, Sobrev.: <30)"
    if ACTIVAR_MACD and 'macd_actual' in stats and pd.notna(stats['macd_actual']):
        indicadores_texto += f"\n• MACD ({MACD_FAST_PERIOD},{MACD_SLOW_PERIOD},{MACD_SIGNAL_PERIOD}): MACD: {stats['macd_actual']:.2f}, Señal: {stats['macd_signal_actual']:.2f}"
    if ACTIVAR_BBANDS and 'bb_upper_actual' in stats and pd.notna(stats['bb_upper_actual']):
        indicadores_texto += f"\n• Bandas de Bollinger ({BBANDS_PERIOD}d): Superior: {stats['bb_upper_actual']:.2f}, Media: {stats['bb_middle_actual']:.2f}, Inferior: {stats['bb_lower_actual']:.2f}"
    if ACTIVAR_VOLUMEN and 'volumen_promedio' in stats and pd.notna(stats['volumen_promedio']):
        indicadores_texto += f"\n• Volumen Promedio: {stats['volumen_promedio']:.2f}"

    if indicadores_texto:
        analisis_texto += f"\n\nINDICADORES TÉCNICOS ADICIONALES:{indicadores_texto}"

    # Mostrar texto
    ax.text(0.05, 0.88, metricas_texto, 
            fontsize=10, ha='left', va='top',
            color='#212121', transform=ax.transAxes,
            family='monospace')
    
    ax.text(0.05, 0.55, analisis_texto, 
            fontsize=10, ha='left', va='top',
            color='#424242', transform=ax.transAxes)
    
    # Espacio para pie de página
    plt.tight_layout()
    plt.subplots_adjust(bottom=0.08) # Ajustado para acomodar el pie de página
    
    agregar_pie_pagina_correcto(fig, numero_pagina, total_paginas, autor_nombre, autor_portfolio)
    
    return fig

def procesar_activos():
    """Procesa todos los activos y genera análisis completo"""
    configurar_matplotlib(TEMA_GRAFICOS)
    
    # Asegurarse de que la carpeta de gráficos exista
    os.makedirs(CARPETA_GRAFICOS, exist_ok=True)
    
    # --- CAMBIO CLAVE: Buscar solo archivos .csv y NO llamar a generar_datos_prueba ---
    archivos = [f for f in os.listdir(CARPETA_DATOS) if f.endswith('.csv')]
    
    if not archivos:
        print(f"ERROR: No se encontraron archivos CSV en la carpeta '{CARPETA_DATOS}'.")
        print("Por favor, asegúrate de tener tus datos en formato .csv en esa carpeta.")
        return {}, {} # Devuelve diccionarios vacíos para indicar que no hay datos para procesar.
    
    tickers = [f.replace('.csv', '') for f in archivos]
    print(f"Procesando {len(tickers)} activos: {', '.join(tickers)}")
    
    # Calcular total de páginas aproximado (portada + 4 páginas por activo)
    # Se añaden 3 páginas más por activo si todos los indicadores están activados (RSI, MACD, Volumen)
    paginas_por_activo = 4 # Precio, Comentario, Retornos, Histograma
    if ACTIVAR_RSI: paginas_por_activo += 1
    if ACTIVAR_MACD: paginas_por_activo += 1
    if ACTIVAR_VOLUMEN: paginas_por_activo += 1

    total_paginas = 1 + (len(tickers) * paginas_por_activo)
    pagina_actual = 1
    
    # Diccionarios para almacenar resultados
    resultados = {}
    dataframes = {}
    
    # Crear PDF
    pdf = PdfPages(ARCHIVO_PDF)
    
    
    # Crear portada
    print("Creando portada del PDF...")
    crear_portada_pdf(pdf, tickers, AUTOR_NOMBRE, AUTOR_PORTFOLIO)
    pagina_actual += 1
    
    # Procesar cada activo
    for archivo in archivos:
        nombre = archivo.replace('.csv', '') # Cambiado de .xlsx/.xls a .csv
        ruta = os.path.join(CARPETA_DATOS, archivo)
        
        print(f"Procesando {nombre}...")
        
        try:
            # Leer el CSV de forma más robusta
            # La primera línea es el encabezado real, la segunda es 'Ticker', la tercera es 'Date,,,,,'.
            # Los datos comienzan en la cuarta línea.
            # Por lo tanto, saltamos las primeras 2 líneas (índice 0 y 1) y usamos la línea 0 (índice 0) como encabezado.
            # La columna de fecha es la primera columna (índice 0).
            df = pd.read_csv(ruta, header=0, skiprows=[1, 2], index_col=0, parse_dates=True)
            
            # Asegurarse de que el índice sea DatetimeIndex
            if not isinstance(df.index, pd.DatetimeIndex):
                df.index = pd.to_datetime(df.index)

            # Ordenar por fecha
            df.sort_index(inplace=True)
            
            # Procesar precios
            # Priorizar 'Adj Close', luego 'Close', luego 'Price'
            if 'Adj Close' in df.columns:
                df.rename(columns={'Adj Close': 'Precio'}, inplace=True)
            elif 'Close' in df.columns:
                df.rename(columns={'Close': 'Precio'}, inplace=True)
            elif 'Price' in df.columns:
                df.rename(columns={'Price': 'Precio'}, inplace=True) # Corregido para usar 'Price' directamente
            else:
                # Si no se encuentra ninguna de las columnas esperadas, intentar con la primera columna de datos numérica
                numeric_cols = df.select_dtypes(include=np.number).columns
                if not numeric_cols.empty:
                    df.rename(columns={numeric_cols[0]: 'Precio'}, inplace=True)
                else:
                    print(f"  Error procesando {nombre}: No se encontró columna de precio adecuada. Columnas disponibles: {df.columns.tolist()}. Saltando este activo.")
                    continue

            # Calcular Medias Móviles
            df[f'SMA_{SMA_CORTA}'] = df['Precio'].rolling(window=SMA_CORTA).mean()
            df[f'SMA_{SMA_MEDIA}'] = df['Precio'].rolling(window=SMA_MEDIA).mean()
            df[f'SMA_{SMA_LARGA}'] = df['Precio'].rolling(window=SMA_LARGA).mean()

            # Calcular Retornos Diarios
            df['Retorno'] = df['Precio'].pct_change().dropna()

            # Calcular RSI
            if ACTIVAR_RSI:
                delta = df['Precio'].diff()
                gain = (delta.where(delta > 0, 0)).rolling(window=RSI_PERIOD).mean()
                loss = (-delta.where(delta < 0, 0)).rolling(window=RSI_PERIOD).mean()
                rs = gain / loss
                df['RSI'] = 100 - (100 / (1 + rs))
            
            # Calcular MACD
            if ACTIVAR_MACD:
                exp1 = df['Precio'].ewm(span=MACD_FAST_PERIOD, adjust=False).mean()
                exp2 = df['Precio'].ewm(span=MACD_SLOW_PERIOD, adjust=False).mean()
                df['MACD'] = exp1 - exp2
                df['MACD_Signal'] = df['MACD'].ewm(span=MACD_SIGNAL_PERIOD, adjust=False).mean()
                df['MACD_Hist'] = df['MACD'] - df['MACD_Signal']

            # Calcular Bandas de Bollinger
            if ACTIVAR_BBANDS:
                df['BB_Middle'] = df['Precio'].rolling(window=BBANDS_PERIOD).mean()
                df['BB_StdDev'] = df['Precio'].rolling(window=BBANDS_PERIOD).std()
                df['BB_Upper'] = df['BB_Middle'] + (df['BB_StdDev'] * BBANDS_DEV)
                df['BB_Lower'] = df['BB_Middle'] - (df['BB_StdDev'] * BBANDS_DEV)

            # Guardar DataFrame procesado
            dataframes[nombre] = df

            # Recopilar estadísticas para el resumen
            precio_actual = df['Precio'].iloc[-1]
            retorno_anual = df['Retorno'].mean() * 252 # 252 días hábiles
            volatilidad_diaria = df['Retorno'].std()
            volatilidad_mensual = volatilidad_diaria * np.sqrt(21) # 21 días hábiles al mes
            
            stats = {
                'precio_actual': precio_actual,
                'retorno_anual': retorno_anual,
                'volatilidad_diaria': volatilidad_diaria,
                'volatilidad_mensual': volatilidad_mensual,
                f'sma_{SMA_CORTA}': df[f'SMA_{SMA_CORTA}'].iloc[-1],
                f'sma_{SMA_MEDIA}': df[f'SMA_{SMA_MEDIA}'].iloc[-1],
                f'sma_{SMA_LARGA}': df[f'SMA_{SMA_LARGA}'].iloc[-1]
            }

            if ACTIVAR_RSI and 'RSI' in df.columns:
                stats['rsi_actual'] = df['RSI'].iloc[-1]
            if ACTIVAR_MACD and 'MACD' in df.columns:
                stats['macd_actual'] = df['MACD'].iloc[-1]
                stats['macd_signal_actual'] = df['MACD_Signal'].iloc[-1]
            if ACTIVAR_BBANDS and 'BB_Upper' in df.columns:
                stats['bb_upper_actual'] = df['BB_Upper'].iloc[-1]
                stats['bb_middle_actual'] = df['BB_Middle'].iloc[-1]
                stats['bb_lower_actual'] = df['BB_Lower'].iloc[-1]
            if ACTIVAR_VOLUMEN and 'Volume' in df.columns:
                stats['volumen_promedio'] = df['Volume'].mean()

            resultados[nombre] = stats

            # Generar gráficos y añadir al PDF
            # Gráfico de Precios
            fig_precio = crear_grafico_precios(df, nombre, ESCALA_PRECIOS, pagina_actual, total_paginas)
            pdf.savefig(fig_precio, bbox_inches='tight')
            plt.close(fig_precio)
            pagina_actual += 1

            # Gráfico de Retornos
            fig_retornos = crear_grafico_retornos(df, nombre, pagina_actual, total_paginas)
            pdf.savefig(fig_retornos, bbox_inches='tight')
            plt.close(fig_retornos)
            pagina_actual += 1

            # Histograma de Retornos
            fig_hist = crear_histograma_retornos(df, nombre, pagina_actual, total_paginas)
            pdf.savefig(fig_hist, bbox_inches='tight')
            plt.close(fig_hist)
            pagina_actual += 1

            # Gráfico de RSI (si está activado)
            if ACTIVAR_RSI:
                fig_rsi = crear_grafico_rsi(df, nombre, pagina_actual, total_paginas)
                pdf.savefig(fig_rsi, bbox_inches='tight')
                plt.close(fig_rsi)
                pagina_actual += 1

            # Gráfico de MACD (si está activado)
            if ACTIVAR_MACD:
                fig_macd = crear_grafico_macd(df, nombre, pagina_actual, total_paginas)
                pdf.savefig(fig_macd, bbox_inches='tight')
                plt.close(fig_macd)
                pagina_actual += 1

            # Gráfico de Volumen (si está activado)
            if ACTIVAR_VOLUMEN:
                fig_volumen = crear_grafico_volumen(df, nombre, pagina_actual, total_paginas)
                pdf.savefig(fig_volumen, bbox_inches='tight')
                plt.close(fig_volumen)
                pagina_actual += 1

            # Página de Comentarios
            fig_comentario = crear_comentario_activo(nombre, stats, df, pagina_actual, total_paginas)
            pdf.savefig(fig_comentario, bbox_inches='tight')
            plt.close(fig_comentario)
            pagina_actual += 1

        except Exception as e:
            print(f"Error procesando {nombre}: {e}")
            continue

    pdf.close()
    print(f"PDF guardado como: {ARCHIVO_PDF}")
    return resultados, dataframes

def generar_excel_resumen(resultados, dataframes):
    """Genera un archivo Excel con el resumen de los resultados y gráficos."""
    if not resultados:
        print("No se pudieron procesar los activos. No se generará el archivo Excel.")
        return

    # Asegurarse de que la carpeta de gráficos exista antes de guardar imágenes temporales
    os.makedirs(CARPETA_GRAFICOS, exist_ok=True)

    # Crear un DataFrame con los resultados para la tabla resumen
    df_resumen = pd.DataFrame.from_dict(resultados, orient='index')
    df_resumen.index.name = 'Activo'
    
    # Formatear columnas para el Excel
    df_resumen['precio_actual'] = df_resumen['precio_actual'].map('${:.2f}'.format)
    df_resumen['retorno_anual'] = df_resumen['retorno_anual'].map('{:.2%}'.format)
    df_resumen['volatilidad_diaria'] = df_resumen['volatilidad_diaria'].map('{:.2%}'.format)
    df_resumen['volatilidad_mensual'] = df_resumen['volatilidad_mensual'].map('{:.2%}'.format)
    
    # Renombrar columnas para el Excel
    df_resumen.rename(columns={
        'precio_actual': 'Precio Actual',
        'retorno_anual': 'Retorno Anual',
        'volatilidad_diaria': 'Volatilidad Diaria',
        'volatilidad_mensual': 'Volatilidad Mensual',
        f'sma_{SMA_CORTA}': f'SMA {SMA_CORTA}d',
        f'sma_{SMA_MEDIA}': f'SMA {SMA_MEDIA}d',
        f'sma_{SMA_LARGA}': f'SMA {SMA_LARGA}d',
        'rsi_actual': 'RSI Actual',
        'macd_actual': 'MACD Actual',
        'macd_signal_actual': 'MACD Señal Actual',
        'bb_upper_actual': 'BB Superior Actual',
        'bb_middle_actual': 'BB Media Actual',
        'bb_lower_actual': 'BB Inferior Actual',
        'volumen_promedio': 'Volumen Promedio'
    }, inplace=True)

    # Crear un nuevo libro de Excel
    writer = pd.ExcelWriter(ARCHIVO_EXCEL, engine='openpyxl')
    df_resumen.to_excel(writer, sheet_name='Resumen', index=True)
    
    # Acceder al workbook y la hoja de resumen
    workbook = writer.book
    sheet_resumen = writer.sheets['Resumen']

    # Ajustar ancho de columnas y aplicar formato a la hoja de resumen
    for column in sheet_resumen.columns:
        max_length = 0
        column_name = column[0].column_letter # Get the column name
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet_resumen.column_dimensions[column_name].width = adjusted_width

    # Estilos para la tabla resumen
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border_style = Side(border_style="thin", color="000000")
    full_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    # Aplicar estilo a los encabezados
    for cell in sheet_resumen["1:1"]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = full_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Aplicar estilo a las celdas de datos
    for row in sheet_resumen.iter_rows(min_row=2):
        for cell in row:
            cell.border = full_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Crear hojas individuales para cada activo con su gráfico
    for nombre, df in dataframes.items():
        sheet = workbook.create_sheet(title=nombre)
        
        # Escribir el DataFrame de datos históricos
        df.to_excel(writer, sheet_name=nombre, index=True)
        
        # Ajustar ancho de columnas para la hoja del activo
        for column in sheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_name].width = adjusted_width

        # Guardar el gráfico de precios como imagen temporal
        temp_img_path = os.path.join(CARPETA_GRAFICOS, f'{nombre}_precio.png')
        fig_precio = crear_grafico_precios(df, nombre, ESCALA_PRECIOS)
        fig_precio.savefig(temp_img_path, bbox_inches='tight')
        plt.close(fig_precio)

        # Insertar la imagen en la hoja de Excel
        img = Image(temp_img_path)
        # Ajustar tamaño de la imagen para que quepa bien
        img.width = 600  # Ancho en píxeles
        img.height = 400 # Alto en píxeles
        
        # Calcular la celda donde insertar la imagen (después de la tabla)
        # Asumiendo que la tabla empieza en A1 y tiene el número de filas de df + 1 (encabezado)
        start_row_img = len(df) + 3 # 2 filas de espacio + 1 fila de encabezado
        sheet.add_image(img, f'A{start_row_img}')
        
        # Eliminar la imagen temporal
        # os.remove(temp_img_path)  # Mover esto al final después de guardar el Excel

    
    writer.close()
    print(f"Excel guardado como: {ARCHIVO_EXCEL}")

    # Eliminar imágenes temporales al final
    for nombre in dataframes.keys():
        temp_img_path = os.path.join(CARPETA_GRAFICOS, f'{nombre}_precio.png')
        if os.path.exists(temp_img_path):
            try:
                os.remove(temp_img_path)
            except Exception as e:
                print(f"No se pudo eliminar la imagen temporal {temp_img_path}: {e}")


# ========================= EJECUCIÓN PRINCIPAL =========================

if __name__ == "__main__":
    resultados_analisis, dataframes_activos = procesar_activos()
    generar_excel_resumen(resultados_analisis, dataframes_activos)


